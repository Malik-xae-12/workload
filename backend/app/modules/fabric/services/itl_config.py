"""ITL Config service – read OTL config and generate/parse watermark Excel."""

import io
import logging

import pyodbc

logger = logging.getLogger(__name__)

DATETIME_TYPES = {"datetime", "datetime2", "date", "datetimeoffset", "timestamp", "timestamptz", "timestamp without time zone", "timestamp with time zone"}

CREATED_HINTS = {"created", "create", "insert", "inserted"}
UPDATED_HINTS = {"updated", "update", "modified", "modify", "changed"}

# ── Colors ────────────────────────────────────────────────────────────────────
CLR_HEADER_SHEET1   = "4472C4"   # blue
CLR_HEADER_SHEET2   = "70AD47"   # green
CLR_ROW_ODD         = "EBF3FB"   # light blue
CLR_ROW_EVEN        = "FFFFFF"   # white
CLR_SUGGESTED       = "FFF2CC"   # soft yellow – suggested field
CLR_HEADER_FONT     = "FFFFFF"
CLR_BORDER          = "BDD7EE"


def _get_odbc_driver() -> str:
    drivers = pyodbc.drivers()
    for d in ("ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server"):
        if d in drivers:
            return d
    raise RuntimeError("No suitable ODBC driver found.")


def _connect(client_id, client_secret, server, database):
    driver = _get_odbc_driver()
    conn_str = (
        f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};"
        "Authentication=ActiveDirectoryServicePrincipal;"
        f"UID={client_id};PWD={client_secret};Encrypt=yes;TrustServerCertificate=no;"
    )
    return pyodbc.connect(conn_str)


def _suggest_watermarks(datetime_cols: list[str]) -> tuple[str | None, str | None]:
    """Return (created_suggestion, updated_suggestion). Auto-fill only when 1 option exists."""
    if not datetime_cols:
        return None, None
    if len(datetime_cols) == 1:
        # Single col: auto-assign to whichever hint matches, else created
        col = datetime_cols[0]
        low = col.lower()
        if any(h in low for h in UPDATED_HINTS):
            return None, col
        return col, None

    created = updated = None
    for col in datetime_cols:
        low = col.lower()
        if not updated and any(h in low for h in UPDATED_HINTS):
            updated = col
        elif not created and any(h in low for h in CREATED_HINTS):
            created = col
    return created, updated


def read_otl_config(client_id, client_secret, server, database, config_schema_name, db_type: str = "") -> list[dict]:
    """Read OneTimeConfigETL joined with SourceInformationSchema for datetime columns."""
    conn = _connect(client_id, client_secret, server, database)
    cursor = conn.cursor()
    is_oracle = "oracle" in (db_type or "").lower()
    is_postgres = "postgres" in (db_type or "").lower() or (db_type or "").lower() in ("pg",)
    try:
        # Get OTL rows
        cursor.execute(f"""
            SELECT Id, SourceSchemaName, SourceTableName
            FROM [{config_schema_name}].[OneTimeConfigETL]
            WHERE IsActive = 1
            ORDER BY Id
        """)
        otl_rows = [{"Id": r[0], "SourceSchemaName": r[1], "SourceTableName": r[2]} for r in cursor.fetchall()]

        # Get datetime columns from SourceInformationSchema.
        # NOTE: Oracle's SourceInformationSchema is populated from all_tab_columns, which has
        # no TABLE_SCHEMA column (it uses OWNER instead), and reports DATA_TYPE with embedded
        # precision (e.g. "TIMESTAMP(6)", "TIMESTAMP(6) WITH TIME ZONE") rather than the bare
        # type name SQL Server/Postgres INFORMATION_SCHEMA.COLUMNS returns.
        # Postgres's own INFORMATION_SCHEMA.COLUMNS returns column names in lowercase
        # (table_schema, table_name, column_name, data_type) since Postgres folds unquoted
        # identifiers to lowercase - unlike SQL Server, whose INFORMATION_SCHEMA.COLUMNS
        # returns the exact upper-case ANSI names. Fabric Warehouse uses case-sensitive
        # collation, so the physical SourceInformationSchema columns for a Postgres
        # connection are lowercase and must be queried/aliased accordingly.
        # The SQL Server path below is unchanged.
        if is_oracle:
            cursor.execute(f"""
                SELECT OWNER AS TABLE_SCHEMA, TABLE_NAME, COLUMN_NAME, DATA_TYPE
                FROM [{config_schema_name}].[SourceInformationSchema]
                WHERE LOWER(DATA_TYPE) = 'date' OR LOWER(DATA_TYPE) LIKE 'timestamp%'
            """)
        elif is_postgres:
            cursor.execute(f"""
                SELECT table_schema AS TABLE_SCHEMA, table_name AS TABLE_NAME,
                       column_name AS COLUMN_NAME, data_type AS DATA_TYPE
                FROM [{config_schema_name}].[SourceInformationSchema]
                WHERE LOWER(data_type) IN ({','.join('?' for _ in DATETIME_TYPES)})
            """, list(DATETIME_TYPES))
        else:
            cursor.execute(f"""
                SELECT TABLE_SCHEMA, TABLE_NAME, COLUMN_NAME, DATA_TYPE
                FROM [{config_schema_name}].[SourceInformationSchema]
                WHERE LOWER(DATA_TYPE) IN ({','.join('?' for _ in DATETIME_TYPES)})
            """, list(DATETIME_TYPES))
        schema_rows = cursor.fetchall()

        # Build lookup: (schema, table) -> [col, ...]
        dt_map: dict[tuple, list[str]] = {}
        for r in schema_rows:
            key = (r[0], r[1])
            dt_map.setdefault(key, []).append(r[2])

        for row in otl_rows:
            key = (row["SourceSchemaName"], row["SourceTableName"])
            dt_cols = dt_map.get(key, [])
            suggested_created, suggested_updated = _suggest_watermarks(dt_cols)
            row["_dt_cols"] = dt_cols
            row["_suggested_created"] = suggested_created
            row["_suggested_updated"] = suggested_updated

        return otl_rows
    finally:
        cursor.close()
        conn.close()


def write_itl_config(client_id, client_secret, server, database, config_schema_name, rows) -> int:
    conn = _connect(client_id, client_secret, server, database)
    cursor = conn.cursor()
    try:
        inserted = 0
        for row in rows:
            cursor.execute(f"""
                IF EXISTS (SELECT 1 FROM [{config_schema_name}].[IncrementalConfigETL] WHERE Id = ?)
                    UPDATE [{config_schema_name}].[IncrementalConfigETL]
                    SET CreatedWatermarkField = ?, UpdatedWatermarkField = ?, IsFullLoad = ?
                    WHERE Id = ?
                ELSE
                    INSERT INTO [{config_schema_name}].[IncrementalConfigETL]
                        (Id, SourceSchemaName, SourceTableName, CreatedWatermarkField, UpdatedWatermarkField, IsFullLoad)
                    VALUES (?, ?, ?, ?, ?, ?)
            """, (
                row.get("Id"),
                row.get("CreatedWaterMarkField") or None,
                row.get("UpdatedWaterMarkField") or None,
                1 if row.get("IsFullLoad") == "FullLoad" else 0,
                row.get("Id"),
                row.get("Id"),
                row.get("SourceSchemaName"),
                row.get("SourceTableName"),
                row.get("CreatedWaterMarkField") or None,
                row.get("UpdatedWaterMarkField") or None,
                1 if row.get("IsFullLoad") == "FullLoad" else 0,
            ))
            inserted += cursor.rowcount
        conn.commit()
        return inserted
    finally:
        cursor.close()
        conn.close()


def ensure_watermark_sp(client_id, client_secret, server, database, config_schema_name, app_mode: str = "fabric") -> None:
    """Standalone entry point: open a connection and create/update UpdateWaterMarkSP
    inside the connection's own Config_<name> schema.

    *app_mode* selects which medallion source layer the SP reads from:
    'finin' reads from LH_Bronze, everything else reads from LH_Silver."""
    conn = _connect(client_id, client_secret, server, database)
    cursor = conn.cursor()
    try:
        _ensure_watermark_sp(cursor, config_schema_name, app_mode)
        conn.commit()
    finally:
        cursor.close()
        conn.close()


def _ensure_watermark_sp(cursor, config_schema_name: str, app_mode: str = "fabric") -> None:
    """Create/update UpdateWaterMarkSP inside the connection's own Config_<name>
    schema (not Log) — schema name is substituted dynamically, never hardcoded."""
    s = config_schema_name
    lakehouse = "LH_Bronze" if app_mode == "finin" else "LH_Silver"
    cursor.execute(f"""
        CREATE OR ALTER PROCEDURE [{s}].[UpdateWaterMarkSP]
        AS
        BEGIN
            SET NOCOUNT ON;

            DECLARE @Sql NVARCHAR(MAX), @Id BIGINT,
                    @SilverSchemaName VARCHAR(128), @SilverTableName VARCHAR(128),
                    @CreatedField VARCHAR(128), @UpdatedField VARCHAR(128),
                    @NewCreatedValue DATETIME2, @NewUpdatedValue DATETIME2;

            DROP TABLE IF EXISTS [WH_MetaData].[{s}].[WaterMarkProcessing];
            CREATE TABLE [WH_MetaData].[{s}].[WaterMarkProcessing] (
                Id BIGINT, SilverSchemaName VARCHAR(128), SilverTableName VARCHAR(128),
                CreatedWaterMarkField VARCHAR(128), UpdatedWaterMarkField VARCHAR(128)
            );

            INSERT INTO [WH_MetaData].[{s}].[WaterMarkProcessing]
                (Id, SilverSchemaName, SilverTableName, CreatedWaterMarkField, UpdatedWaterMarkField)
            SELECT Id, SilverSchemaName, SilverTableName, CreatedWaterMarkField, UpdatedWaterMarkField
            FROM [WH_MetaData].[{s}].[IncrementalConfigETL]
            WHERE IsFullLoad = 0 AND IsActive = 1
              AND (CreatedWaterMarkField IS NOT NULL AND CreatedWaterMarkField <> '1')
               OR (UpdatedWaterMarkField IS NOT NULL AND UpdatedWaterMarkField <> '1');

            WHILE EXISTS (SELECT 1 FROM [WH_MetaData].[{s}].[WaterMarkProcessing])
            BEGIN
                SELECT TOP 1 @Id = Id, @SilverSchemaName = SilverSchemaName, @SilverTableName = SilverTableName,
                    @CreatedField = CreatedWaterMarkField, @UpdatedField = UpdatedWaterMarkField
                FROM [WH_MetaData].[{s}].[WaterMarkProcessing];

                SET @NewCreatedValue = NULL;
                SET @NewUpdatedValue = NULL;

                BEGIN TRY
                    IF @CreatedField IS NOT NULL AND @CreatedField <> '1' AND ISNULL(@UpdatedField, '') IN ('', '1')
                    BEGIN
                        SET @Sql = N'SELECT @NewCreatedValue = MAX(' + QUOTENAME(@CreatedField) + N')
                                     FROM {lakehouse}.' + QUOTENAME(@SilverSchemaName) + N'.' + QUOTENAME(@SilverTableName);
                        EXEC sp_executesql @Sql, N'@NewCreatedValue DATETIME2 OUTPUT', @NewCreatedValue OUTPUT;
                    END
                    ELSE IF @UpdatedField IS NOT NULL AND @UpdatedField <> '1' AND ISNULL(@CreatedField, '') IN ('', '1')
                    BEGIN
                        SET @Sql = N'SELECT @NewUpdatedValue = MAX(' + QUOTENAME(@UpdatedField) + N')
                                     FROM {lakehouse}.' + QUOTENAME(@SilverSchemaName) + N'.' + QUOTENAME(@SilverTableName);
                        EXEC sp_executesql @Sql, N'@NewUpdatedValue DATETIME2 OUTPUT', @NewUpdatedValue OUTPUT;
                    END
                    ELSE IF @CreatedField IS NOT NULL AND @CreatedField <> '1' AND @UpdatedField IS NOT NULL AND @UpdatedField <> '1'
                    BEGIN
                        SET @Sql = N'SELECT @NewCreatedValue = MAX(' + QUOTENAME(@CreatedField) + N'), @NewUpdatedValue = MAX(' + QUOTENAME(@UpdatedField) + N')
                                     FROM {lakehouse}.' + QUOTENAME(@SilverSchemaName) + N'.' + QUOTENAME(@SilverTableName);
                        EXEC sp_executesql @Sql, N'@NewCreatedValue DATETIME2 OUTPUT, @NewUpdatedValue DATETIME2 OUTPUT',
                            @NewCreatedValue OUTPUT, @NewUpdatedValue OUTPUT;
                    END

                    UPDATE [WH_MetaData].[{s}].[IncrementalConfigETL]
                    SET CreatedWaterMarkValue = @NewCreatedValue, UpdatedWaterMarkValue = @NewUpdatedValue,
                        LastModifiedDate = SYSUTCDATETIME()
                    WHERE Id = @Id;
                END TRY
                BEGIN CATCH
                    PRINT 'FAILED - Id ' + CAST(@Id AS VARCHAR(10)) + ' | Error: ' + ERROR_MESSAGE();
                END CATCH

                DELETE FROM [WH_MetaData].[{s}].[WaterMarkProcessing] WHERE Id = @Id;
            END

            DROP TABLE IF EXISTS [WH_MetaData].[{s}].[WaterMarkProcessing];
        END
    """)


def _apply_border(ws, cell):
    from openpyxl.styles import Border, Side
    thin = Side(style="thin", color=CLR_BORDER)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_header(cell, hex_color):
    from openpyxl.styles import PatternFill, Font, Alignment
    cell.fill = PatternFill("solid", fgColor=hex_color)
    cell.font = Font(bold=True, color=CLR_HEADER_FONT, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _apply_border(None, cell)


def _style_row(cell, row_idx, highlight=False):
    from openpyxl.styles import PatternFill, Alignment
    if highlight:
        color = CLR_SUGGESTED
    else:
        color = CLR_ROW_ODD if row_idx % 2 == 1 else CLR_ROW_EVEN
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    _apply_border(None, cell)


def generate_itl_excel(rows: list[dict]) -> bytes:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()

    # ── Sheet 1: Table Selection ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Table_Selection"
    ws1.row_dimensions[1].height = 30

    # A=Id B=Schema C=Table D=IsFullLoad E=SuggestedCreated F=SuggestedUpdated
    headers1 = ["Id", "SourceSchemaName", "SourceTableName", "IsFullLoad",
                "SuggestedCreatedWaterMarkField", "SuggestedUpdatedWaterMarkField"]
    col_widths1 = [6, 22, 28, 22, 32, 32]

    for ci, (h, w) in enumerate(zip(headers1, col_widths1), 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        _style_header(cell, CLR_HEADER_SHEET1)
        ws1.column_dimensions[cell.column_letter].width = w

    dv_load = DataValidation(type="list", formula1='"FullLoad,IncrementalLoad"', allow_blank=True)
    dv_load.sqref = f"D2:D{len(rows) + 1}"
    ws1.add_data_validation(dv_load)

    for ri, row in enumerate(rows, 2):
        dt_cols = row.get("_dt_cols", [])
        sc = row.get("_suggested_created")
        su = row.get("_suggested_updated")
        is_full = "FullLoad" if not dt_cols else "IncrementalLoad"

        vals = [row.get("Id"), row.get("SourceSchemaName"), row.get("SourceTableName"),
                is_full, sc or "", su or ""]
        for ci, val in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            _style_row(cell, ri - 1, highlight=(ci in (5, 6) and bool(val)))

        if dt_cols:
            formula = '"' + ",".join(dt_cols) + '"'
            for col_letter in ("E", "F"):
                dv = DataValidation(type="list", formula1=formula, allow_blank=True)
                dv.sqref = f"{col_letter}{ri}"
                ws1.add_data_validation(dv)

    # ── Sheet 2: ITL_Config ───────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="ITL_Config")
    ws2.row_dimensions[1].height = 30

    headers2 = ["Id", "SourceSchemaName", "SourceTableName", "CreatedWaterMarkField", "UpdatedWaterMarkField"]
    col_widths2 = [6, 22, 28, 28, 28]

    for ci, (h, w) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        _style_header(cell, CLR_HEADER_SHEET2)
        ws2.column_dimensions[cell.column_letter].width = w

    for ri, row in enumerate(rows, 2):
        # Cols A-C: static values
        for ci, val in enumerate([row.get("Id"), row.get("SourceSchemaName"), row.get("SourceTableName")], 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            _style_row(cell, ri - 1)
        # Cols D-E: formula → Table_Selection!E{ri} / F{ri} so edits auto-reflect
        for ci, src_col in ((4, "E"), (5, "F")):
            cell = ws2.cell(row=ri, column=ci, value=f"=Table_Selection!{src_col}{ri}")
            _style_row(cell, ri - 1, highlight=True)

    # Freeze header rows
    ws1.freeze_panes = "A2"
    ws2.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def parse_itl_excel(file_bytes: bytes) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    def _sheet_to_map(ws):
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            return {}, []
        hmap = {str(h).strip(): idx for idx, h in enumerate(headers) if h}
        return hmap, list(rows_iter)

    # Sheet 1: Table_Selection (IsFullLoad + user-chosen watermarks in E/F)
    ws1 = wb["Table_Selection"] if "Table_Selection" in wb.sheetnames else None
    s1_map, s1_rows = _sheet_to_map(ws1) if ws1 else ({}, [])

    # Sheet 2: ITL_Config (CreatedWaterMarkField, UpdatedWaterMarkField)
    ws2 = wb["ITL_Config"] if "ITL_Config" in wb.sheetnames else wb.active
    s2_map, s2_rows = _sheet_to_map(ws2)

    required = ["Id", "SourceSchemaName", "SourceTableName"]
    for r in required:
        if r not in s2_map:
            raise ValueError(f"Missing required column in ITL_Config sheet: {r}")

    # Build sheet1 lookup by Id
    s1_by_id = {}
    if s1_map and s1_rows:
        for row in s1_rows:
            if not row or all(c is None for c in row):
                continue
            def _g1(key):
                idx = s1_map.get(key, -1)
                return row[idx] if 0 <= idx < len(row) else None
            rid = _g1("Id")
            if rid is not None:
                s1_by_id[rid] = {
                    "IsFullLoad": _g1("IsFullLoad"),
                    "SuggestedCreatedWaterMarkField": _g1("SuggestedCreatedWaterMarkField"),
                    "SuggestedUpdatedWaterMarkField": _g1("SuggestedUpdatedWaterMarkField"),
                }

    results = []
    for row in s2_rows:
        if not row or all(c is None for c in row):
            continue
        def _get(key):
            idx = s2_map.get(key, -1)
            if idx < 0 or idx >= len(row):
                return None
            v = row[idx]
            return None if v in (None, "", "NULL") else v

        rid = _get("Id")
        s1 = s1_by_id.get(rid, {})

        # Prefer sheet2 watermarks; fall back to sheet1 E/F selections
        created = _get("CreatedWaterMarkField") or s1.get("SuggestedCreatedWaterMarkField") or None
        updated = _get("UpdatedWaterMarkField") or s1.get("SuggestedUpdatedWaterMarkField") or None

        is_full_load = s1.get("IsFullLoad") or ("FullLoad" if not created and not updated else "IncrementalLoad")

        results.append({
            "Id": rid,
            "SourceSchemaName": _get("SourceSchemaName"),
            "SourceTableName": _get("SourceTableName"),
            "IsFullLoad": is_full_load,
            "CreatedWaterMarkField": created,
            "UpdatedWaterMarkField": updated,
        })

    return results