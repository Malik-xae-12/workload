"""Build a Fabric Semantic Model on top of WH_Gold.

Flow:
  1. parse_semantic_excel()      – read the user-uploaded Excel: which
     WH_Gold tables to include, the relationships between them, and the
     measures to create (schema/table names, relationships and measures all
     come from the workbook; column lists are read live from WH_Gold so the
     model always reflects the warehouse's actual columns/types).
  2. fetch_table_columns()       – pyodbc against WH_Gold's SQL endpoint,
     one INFORMATION_SCHEMA.COLUMNS query per selected table.
  3. build_model_bim()           – assemble a TMSL ("model.bim") definition:
     a DirectQuery expression pointing at WH_Gold, one table per selected
     table (with its live columns), the relationships and measures from the
     Excel.
  4. create_semantic_model()     – POST the definition to the Fabric REST
     API (Items - Create Semantic Model); this is a long-running operation,
     so poll the Location URL the same way notebook.py / pipeline.py do.

The per-project WH_Gold/workspace resolution (via _resolve_gold_and_meta())
and the Gold-folder placement live in service.py, which calls into this
module. HARDCODED_WORKSPACE_ID/HARDCODED_GOLD_SERVER/HARDCODED_GOLD_DATABASE
below are kept only as a fallback for ad-hoc/manual testing of this module
in isolation — the real request path no longer uses them.
"""

from __future__ import annotations

import base64
import io
import json
import logging
import time
import uuid

import httpx
import pyodbc

from app.modules.fabric.services.auth import FABRIC_API_BASE

logger = logging.getLogger(__name__)

_TIMEOUT = httpx.Timeout(120.0, connect=10.0)
_LRO_POLL_INTERVAL = 3   # seconds between LRO status checks
_LRO_MAX_POLLS = 60      # ~3 minutes total

# ── HARD-CODED test target (see module docstring) ──────────────────────
# TODO: replace with the calling project's own Medallion gold_item_id /
# workspace_id, the same way _resolve_gold_and_meta() resolves WH_Gold for
# the Master SP step, once this has been validated end-to-end.
HARDCODED_WORKSPACE_ID = "9da6fcea-b918-44df-95ce-b061e8e4e6c2"
HARDCODED_GOLD_SERVER = (
    "2ybikadzh7yenkzxsnkoh7ua74-5l6knhiyxhpujfoowbq6rzhgyi.datawarehouse.fabric.microsoft.com"
)
HARDCODED_GOLD_DATABASE = "WH_Gold"


def _headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


def _get_odbc_driver() -> str:
    drivers = pyodbc.drivers()
    for d in ("ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server"):
        if d in drivers:
            return d
    raise RuntimeError("No suitable ODBC driver found.")


def _connect(client_id: str, client_secret: str, server: str, database: str, timeout: int | None = None):
    driver = _get_odbc_driver()
    conn_str = (
        f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};"
        "Authentication=ActiveDirectoryServicePrincipal;"
        f"UID={client_id};PWD={client_secret};Encrypt=yes;TrustServerCertificate=no;"
    )
    kwargs = {"autocommit": True}
    if timeout is not None:
        kwargs["timeout"] = timeout
    return pyodbc.connect(conn_str, **kwargs)


# ── 1. Excel parsing ────────────────────────────────────────────────────
#
# Only two facts are actually required to build a semantic model table:
# which SCHEMA it lives in, and its TABLE NAME. Everything else the model
# needs (columns/types) is read live from WH_Gold in fetch_table_columns(),
# and everything else the TMSL needs (lineage tags, partition mode/M
# expression) is generated fresh in build_model_bim(). So the parser below
# is deliberately tolerant of extra "export noise" columns (Description,
# Partition Mode, Expression Source, Lineage Tag, Source Lineage Tag, ...)
# that a workbook copied out of an existing model may carry — it reads only
# what it needs and ignores the rest.
#
# Expected workbook layout:
#   A "Tables" sheet (or, if absent, the first sheet that has a table-name
#     column) with a table-name column (either header works):
#       TableName | Source Entity
#     and, optionally, a schema column (either header works):
#       SchemaName | Source Schema        (defaults to "dbo" if omitted)
#   Sheet "Relationships" (optional) FromSchema | FromTable | FromColumn |
#       ToSchema | ToTable | ToColumn | Cardinality | CrossFilterDirection |
#       IsActive (optional, defaults to true — set false on the redundant
#       side of any relationship cycle, same as Power BI/Tabular requires,
#       or the Fabric provisioning API rejects the model with an
#       "ambiguous paths" error)
#   Sheet "Measures" (optional)      SchemaName | TableName | MeasureName |
#       Expression | FormatString
# Column order doesn't matter (matched by header name); schema columns,
# FormatString, and CrossFilterDirection are optional and default sensibly.

_TABLE_NAME_HEADERS = ("TableName", "Source Entity")
_SCHEMA_HEADERS = ("SchemaName", "Source Schema")


def _sheet_rows(ws) -> tuple[dict[str, int], list[tuple]]:
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return {}, []
    hmap = {str(h).strip(): idx for idx, h in enumerate(headers) if h}
    return hmap, [r for r in rows_iter if r and any(c is not None for c in r)]


def _get_first(hmap: dict, row: tuple, *keys: str):
    """Return the first non-empty value among the given header aliases."""
    for key in keys:
        idx = hmap.get(key, -1)
        if idx < 0 or idx >= len(row):
            continue
        v = row[idx]
        if isinstance(v, str):
            v = v.strip()
        if v not in (None, "", "NULL"):
            return v
    return None


def _find_tables_sheet(wb) -> tuple:
    """Return (worksheet, header_map, data_rows) for the sheet that holds
    the table list. Prefers a sheet literally named "Tables"; otherwise
    falls back to the first sheet that has a recognizable table-name
    column, so exports with different sheet names (e.g. a Fabric model
    documentation export) still work without renaming anything."""
    if "Tables" in wb.sheetnames:
        hmap, rows = _sheet_rows(wb["Tables"])
        return wb["Tables"], hmap, rows

    for name in wb.sheetnames:
        hmap, rows = _sheet_rows(wb[name])
        if any(h in hmap for h in _TABLE_NAME_HEADERS):
            return wb[name], hmap, rows

    raise ValueError(
        "No sheet found with a table-name column "
        f"(expected a 'Tables' sheet, or a column named one of: {', '.join(_TABLE_NAME_HEADERS)})"
    )


def parse_semantic_excel(file_bytes: bytes) -> dict:
    """Parse the uploaded workbook into {tables, relationships, measures}."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    def _get(hmap: dict, row: tuple, key: str):
        return _get_first(hmap, row, key)

    # Tables — only TableName/Source Entity + optional SchemaName/Source
    # Schema are read; any other columns present in the sheet are ignored.
    _, t_map, t_rows = _find_tables_sheet(wb)

    tables = []
    seen = set()
    for row in t_rows:
        table_name = _get_first(t_map, row, *_TABLE_NAME_HEADERS)
        if not table_name:
            continue
        schema_name = _get_first(t_map, row, *_SCHEMA_HEADERS) or "dbo"
        key = (str(schema_name), str(table_name))
        if key in seen:
            continue  # de-dupe, in case the export has repeated rows
        seen.add(key)
        tables.append({"schema_name": str(schema_name), "table_name": str(table_name)})

    # Relationships (optional sheet)
    relationships = []
    if "Relationships" in wb.sheetnames:
        r_map, r_rows = _sheet_rows(wb["Relationships"])
        required = ["FromTable", "FromColumn", "ToTable", "ToColumn"]
        missing = [c for c in required if c not in r_map]
        if missing:
            raise ValueError(f"Relationships sheet missing column(s): {', '.join(missing)}")
        for row in r_rows:
            from_table = _get(r_map, row, "FromTable")
            to_table = _get(r_map, row, "ToTable")
            from_col = _get(r_map, row, "FromColumn")
            to_col = _get(r_map, row, "ToColumn")
            if not (from_table and to_table and from_col and to_col):
                continue
            relationships.append({
                "from_schema": str(_get(r_map, row, "FromSchema") or "dbo"),
                "from_table": str(from_table),
                "from_column": str(from_col),
                "to_schema": str(_get(r_map, row, "ToSchema") or "dbo"),
                "to_table": str(to_table),
                "to_column": str(to_col),
                "cardinality": str(_get(r_map, row, "Cardinality") or "many-to-one").lower(),
                "cross_filter_direction": str(
                    _get(r_map, row, "CrossFilterDirection") or "single"
                ).lower(),
                # Defaults to active. Tabular models reject a definition
                # where two tables are connected by more than one active
                # filter path ("ambiguous paths" error from the Fabric
                # provisioning API) — an Excel exported from an existing
                # model marks the redundant side of any such cycle
                # inactive for exactly this reason, so that flag has to
                # survive the round-trip rather than defaulting every
                # relationship to active.
                "is_active": str(_get(r_map, row, "IsActive") or "true").strip().lower()
                not in ("false", "0", "no"),
            })

    # Measures (optional sheet)
    measures = []
    if "Measures" in wb.sheetnames:
        m_map, m_rows = _sheet_rows(wb["Measures"])
        required = ["TableName", "MeasureName", "Expression"]
        missing = [c for c in required if c not in m_map]
        if missing:
            raise ValueError(f"Measures sheet missing column(s): {', '.join(missing)}")
        for row in m_rows:
            table_name = _get(m_map, row, "TableName")
            measure_name = _get(m_map, row, "MeasureName")
            expression = _get(m_map, row, "Expression")
            if not (table_name and measure_name and expression):
                continue
            measures.append({
                "schema_name": str(_get(m_map, row, "SchemaName") or "dbo"),
                "table_name": str(table_name),
                "measure_name": str(measure_name),
                "expression": str(expression),
                "format_string": _get(m_map, row, "FormatString"),
            })

    if not tables:
        raise ValueError("No tables found in the Tables sheet")

    return {"tables": tables, "relationships": relationships, "measures": measures}


# ── 2. WH_Gold schema introspection ─────────────────────────────────────

_SQL_TO_TMSL_TYPE = {
    "bit": "boolean",
    "tinyint": "int64", "smallint": "int64", "int": "int64", "bigint": "int64",
    "decimal": "decimal", "numeric": "decimal", "money": "decimal", "smallmoney": "decimal",
    "float": "double", "real": "double",
    "date": "dateTime", "datetime": "dateTime", "datetime2": "dateTime",
    "smalldatetime": "dateTime", "datetimeoffset": "dateTime", "time": "dateTime",
    "char": "string", "varchar": "string", "nchar": "string", "nvarchar": "string", "text": "string", "ntext": "string",
    "uniqueidentifier": "string",
}


def _tables_visible_in_catalog(cursor, schema_table_pairs: list[tuple[str, str]]) -> set[tuple[str, str]]:
    """Which of the given (schema, table) pairs exist per sys.tables/sys.views.
    sys.tables is basic catalog metadata — visible to any identity that can
    connect at all, regardless of per-object SELECT grants — so this tells
    us whether a table is genuinely absent vs. present-but-unreadable via
    INFORMATION_SCHEMA.COLUMNS (which IS filtered by SELECT permission)."""
    if not schema_table_pairs:
        return set()
    cursor.execute(
        """
        SELECT s.name, t.name FROM sys.tables t JOIN sys.schemas s ON t.schema_id = s.schema_id
        UNION
        SELECT s.name, v.name FROM sys.views v JOIN sys.schemas s ON v.schema_id = s.schema_id
        """
    )
    all_objects = {(r[0], r[1]) for r in cursor.fetchall()}
    wanted = set(schema_table_pairs)
    return all_objects & wanted


def fetch_table_columns(
    client_id: str, client_secret: str, server: str, database: str,
    tables: list[dict],
) -> dict[str, list[dict]]:
    """For each {schema_name, table_name} in `tables`, read its live columns
    from WH_Gold. Returns {"schema.table": [{"name":..,"tmsl_type":..}, ...]}.
    Raises if a table from the Excel doesn't actually exist in WH_Gold —
    better to fail loudly here than to silently build a model missing it."""
    conn = _connect(client_id, client_secret, server, database, timeout=30)
    try:
        cursor = conn.cursor()
        result: dict[str, list[dict]] = {}
        missing: list[str] = []
        for t in tables:
            schema, name = t["schema_name"], t["table_name"]
            cursor.execute(
                """
                SELECT COLUMN_NAME, DATA_TYPE
                FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
                ORDER BY ORDINAL_POSITION
                """,
                schema, name,
            )
            cols = [
                {"name": r[0], "tmsl_type": _SQL_TO_TMSL_TYPE.get((r[1] or "").lower(), "string")}
                for r in cursor.fetchall()
            ]
            if not cols:
                missing.append(f"{schema}.{name}")
                continue
            result[f"{schema}.{name}"] = cols
        if missing:
            # Distinguish "genuinely doesn't exist" from "exists but this
            # identity has no SELECT grant on it" — INFORMATION_SCHEMA.COLUMNS
            # is filtered by per-object SELECT permission, sys.tables/sys.views
            # is not, so a mismatch between the two pinpoints a permissions
            # problem instead of a naming typo in the Tables sheet.
            missing_pairs = [
                (t["schema_name"], t["table_name"]) for t in tables
                if f"{t['schema_name']}.{t['table_name']}" in missing
            ]
            visible = _tables_visible_in_catalog(cursor, missing_pairs)
            no_permission = sorted(f"{s}.{n}" for s, n in visible)
            truly_missing = sorted(m for m in missing if m not in no_permission)

            parts = []
            if no_permission:
                parts.append(
                    f"exist in {database} but this service principal has no SELECT "
                    f"permission on them: {', '.join(no_permission)} "
                    f"(grant it, e.g. GRANT SELECT ON SCHEMA::<schema> TO [<sp-name>])"
                )
            if truly_missing:
                parts.append(
                    f"not found in {database} at all: {', '.join(truly_missing)} "
                    "(check SchemaName/TableName in the Tables sheet)"
                )
            raise ValueError("Table(s) " + "; ".join(parts))
        return result
    finally:
        conn.close()


# ── 2b. Relationship & measure auto-detection from WH_Gold ─────────────
#
# When the uploaded Excel has no Relationships/Measures sheet (or the sheet
# is empty — this is the normal case for an export like the FinIn Analytics
# Model Analysis workbook, which only lists tables), the model should still
# come out with real relationships and measures instead of none at all.
# Both are derived straight from WH_Gold rather than guessed blindly:
#
#   Relationships — first from WH_Gold's own declared FOREIGN KEY
#   constraints (Fabric Warehouses support NOT ENFORCED FKs specifically so
#   modeling tools can read them as ground truth). For any selected table
#   that still isn't related to anything after that, a conservative naming-
#   convention fallback looks for an exact key-column match between exactly
#   two tables (e.g. a "DimAccountKey" column on both a fact table and
#   ims.DimAccount) — it only fires when the match is unambiguous.
#
#   Measures — a "measures control table" convention: a selected table
#   whose name contains "measure" (e.g. ims.AllMeasures) isn't really a
#   data table, it's a row-per-measure catalog (MeasureName / Expression /
#   FormatString / target table, several header spellings accepted). Its
#   ROWS are read and turned into DAX measures attached to whichever
#   selected table each row names; the control table itself is then left
#   out of the model's regular tables (it's metadata, not analytical data).

_KEY_COLUMN_SUFFIXES = ("Key", "ID", "Id", "Code")

_MEASURE_TABLE_HINTS = ("measure",)
_MEASURE_NAME_HEADERS = ("MeasureName", "Measure Name", "Name")
_MEASURE_EXPR_HEADERS = ("Expression", "DAX Expression", "DAXExpression", "Formula", "DAX")
_MEASURE_FORMAT_HEADERS = ("FormatString", "Format String", "Format")
_MEASURE_TARGET_TABLE_HEADERS = (
    "TableName", "Target Table", "TargetTable", "HomeTable", "Home Table", "DisplayTable",
)


def fetch_foreign_key_relationships(
    client_id: str, client_secret: str, server: str, database: str,
    tables: list[dict],
) -> list[dict]:
    """Read WH_Gold's own FK constraints and keep only the ones where both
    sides are among the selected tables. Returns the same shape as
    Excel-parsed relationships, so it can be fed straight into
    build_model_bim()."""
    selected = {(t["schema_name"], t["table_name"]) for t in tables}
    conn = _connect(client_id, client_secret, server, database, timeout=30)
    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT
                sch1.name, tab1.name, col1.name,
                sch2.name, tab2.name, col2.name
            FROM sys.foreign_keys fk
            JOIN sys.foreign_key_columns fkc
                ON fkc.constraint_object_id = fk.object_id
            JOIN sys.tables tab1 ON fkc.parent_object_id = tab1.object_id
            JOIN sys.schemas sch1 ON tab1.schema_id = sch1.schema_id
            JOIN sys.columns col1
                ON fkc.parent_object_id = col1.object_id
               AND fkc.parent_column_id = col1.column_id
            JOIN sys.tables tab2 ON fkc.referenced_object_id = tab2.object_id
            JOIN sys.schemas sch2 ON tab2.schema_id = sch2.schema_id
            JOIN sys.columns col2
                ON fkc.referenced_object_id = col2.object_id
               AND fkc.referenced_column_id = col2.column_id
            """
        )
        relationships = []
        for from_schema, from_table, from_col, to_schema, to_table, to_col in cursor.fetchall():
            if (from_schema, from_table) not in selected or (to_schema, to_table) not in selected:
                continue  # FK involves a table that isn't part of this model — skip it
            relationships.append({
                "from_schema": from_schema, "from_table": from_table, "from_column": from_col,
                "to_schema": to_schema, "to_table": to_table, "to_column": to_col,
                "cardinality": "many-to-one",
                "cross_filter_direction": "single",
                "is_active": True,
            })
        return relationships
    finally:
        conn.close()


def infer_relationships_by_key_convention(
    tables_columns: dict[str, list[dict]],
    tables: list[dict],
    already_related: set[tuple] | None = None,
) -> list[dict]:
    """Best-effort fallback for table pairs with no declared FK: if a
    column name (e.g. "DimAccountKey") appears in exactly two selected
    tables and looks like a key (ends in Key/ID/Code), treat it as a
    many-to-one relationship into whichever of the two tables its name
    prefixes (e.g. "DimAccountKey" -> table "DimAccount" is the "one"
    side). Deliberately conservative — skips anything ambiguous (shared by
    more than two tables, or matching neither table's name) rather than
    inventing a wrong relationship."""
    already_related = already_related or set()

    owners: dict[str, list[tuple]] = {}
    for t in tables:
        key = f"{t['schema_name']}.{t['table_name']}"
        for c in tables_columns.get(key, []):
            if c["name"].endswith(_KEY_COLUMN_SUFFIXES):
                owners.setdefault(c["name"], []).append((t["schema_name"], t["table_name"]))

    relationships = []
    for col_name, holder_tables in owners.items():
        if len(holder_tables) != 2:
            continue  # not a two-table key match — too ambiguous to guess
        (s1, t1), (s2, t2) = holder_tables
        if (s1, t1, s2, t2) in already_related or (s2, t2, s1, t1) in already_related:
            continue
        if col_name.lower().startswith(t1.lower()):
            one_side, many_side = (s1, t1), (s2, t2)
        elif col_name.lower().startswith(t2.lower()):
            one_side, many_side = (s2, t2), (s1, t1)
        else:
            continue  # neither table name prefixes the column — too uncertain
        relationships.append({
            "from_schema": many_side[0], "from_table": many_side[1], "from_column": col_name,
            "to_schema": one_side[0], "to_table": one_side[1], "to_column": col_name,
            "cardinality": "many-to-one",
            "cross_filter_direction": "single",
            "is_active": True,
        })
    return relationships


def _select_all_rows(cursor, database: str, schema: str, name: str):
    """SELECT * from a table, tolerating the couple of ways Fabric SQL
    endpoints can reject a name that INFORMATION_SCHEMA still reports:
    tries the normal 2-part [schema].[table] name first, then falls back
    to the fully qualified 3-part [database].[schema].[table] name (some
    Fabric Warehouse SQL analytics endpoints need this for certain
    objects). Returns (rows, column_names), or (None, None) if neither
    form is queryable — callers should treat that as "leave this table
    alone" rather than fail the whole build over it."""
    for qualified_name in (f"[{schema}].[{name}]", f"[{database}].[{schema}].[{name}]"):
        try:
            cursor.execute(f"SELECT * FROM {qualified_name}")
            col_names = [d[0] for d in cursor.description]
            return cursor.fetchall(), col_names
        except pyodbc.Error as e:
            logger.warning(f"SELECT * FROM {qualified_name} failed: {e}")
            continue
    return None, None


def fetch_measures_from_gold(
    client_id: str, client_secret: str, server: str, database: str,
    tables: list[dict],
) -> tuple[list[dict], set[tuple]]:
    """If one of the selected tables is a measures-control table (name
    contains "measure", e.g. ims.AllMeasures), read its ROWS — not its
    column schema — and turn each row into a DAX measure attached to
    whichever selected table its target-table column names. Returns
    (measures, control_table_keys); control_table_keys tells the caller
    which selected tables were actually measure catalogs rather than real
    data tables, so they can be left out of the model's regular tables."""
    candidates = [
        t for t in tables
        if any(h in t["table_name"].lower() for h in _MEASURE_TABLE_HINTS)
    ]
    if not candidates:
        return [], set()

    valid_targets = {t["table_name"]: t["schema_name"] for t in tables}
    measures: list[dict] = []
    control_keys: set[tuple] = set()

    conn = _connect(client_id, client_secret, server, database, timeout=30)
    try:
        cursor = conn.cursor()
        for t in candidates:
            schema, name = t["schema_name"], t["table_name"]
            rows, col_names = _select_all_rows(cursor, database, schema, name)
            if rows is None:
                # Couldn't query this table at all (doesn't exist as a
                # queryable object, no permission, etc.) — leave it as a
                # normal data table rather than failing the whole build;
                # fetch_table_columns() will surface a clearer error later
                # if it turns out not to exist as a table either.
                logger.warning(
                    f"Could not read rows from {schema}.{name} to check for measures "
                    "(see warning above for the underlying SQL error) — treating it as "
                    "a normal data table instead."
                )
                continue
            hmap = {c: i for i, c in enumerate(col_names)}

            # Doesn't actually look like a measure catalog (no recognizable
            # name/expression columns) — leave it as a normal data table.
            if not (any(h in hmap for h in _MEASURE_NAME_HEADERS)
                    and any(h in hmap for h in _MEASURE_EXPR_HEADERS)):
                continue

            def _cell(row, *keys):
                for k in keys:
                    idx = hmap.get(k)
                    if idx is not None and row[idx] not in (None, ""):
                        return row[idx]
                return None

            control_keys.add((schema, name))
            for row in rows:
                measure_name = _cell(row, *_MEASURE_NAME_HEADERS)
                expression = _cell(row, *_MEASURE_EXPR_HEADERS)
                if not measure_name or not expression:
                    continue
                target_table = _cell(row, *_MEASURE_TARGET_TABLE_HEADERS) or name
                target_schema = valid_targets.get(str(target_table))
                if target_schema is None:
                    continue  # measure points at a table not in this model — skip rather than guess
                measures.append({
                    "schema_name": target_schema,
                    "table_name": str(target_table),
                    "measure_name": str(measure_name),
                    "expression": str(expression),
                    "format_string": _cell(row, *_MEASURE_FORMAT_HEADERS),
                })
        return measures, control_keys
    finally:
        conn.close()


def auto_detect_relationships_and_measures(
    client_id: str, client_secret: str, server: str, database: str,
    tables: list[dict],
    existing_relationships: list[dict],
    existing_measures: list[dict],
) -> tuple[list[dict], list[dict], list[dict]]:
    """Fill in relationships/measures from WH_Gold when the Excel didn't
    provide them explicitly (empty Relationships/Measures sheets, or no
    such sheets at all). Excel-provided relationships/measures are treated
    as an explicit override and always win over auto-detection.

    Returns (tables, relationships, measures) — `tables` has any
    measures-control table(s) like ims.AllMeasures removed, since those
    are measure metadata, not analytical data, and shouldn't show up as a
    regular table in the model.
    """
    measures = list(existing_measures)
    measures_from_gold, control_keys = fetch_measures_from_gold(
        client_id, client_secret, server, database, tables
    )
    if not measures:
        measures = measures_from_gold

    remaining_tables = [
        t for t in tables
        if (t["schema_name"], t["table_name"]) not in control_keys
    ]

    relationships = list(existing_relationships)
    if not relationships:
        relationships = fetch_foreign_key_relationships(
            client_id, client_secret, server, database, remaining_tables
        )
        related = {r["from_table"] for r in relationships} | {r["to_table"] for r in relationships}
        if len(related) < len(remaining_tables):
            # Some tables still have no relationship at all — try the
            # naming-convention fallback for just those gaps.
            tables_columns = fetch_table_columns(
                client_id, client_secret, server, database, remaining_tables
            )
            already = {
                (r["from_schema"], r["from_table"], r["to_schema"], r["to_table"])
                for r in relationships
            }
            relationships += infer_relationships_by_key_convention(
                tables_columns, remaining_tables, already_related=already
            )

    return remaining_tables, relationships, measures


# ── 3. TMSL (model.bim) construction ────────────────────────────────────

_CARDINALITY_MAP = {
    # (fromCardinality, toCardinality) — "from" is the Excel's FromTable side.
    "many-to-one": ("many", "one"),
    "one-to-many": ("one", "many"),
    "one-to-one": ("one", "one"),
    "many-to-many": ("many", "many"),
}


class _UnionFind:
    """Tracks which tables are already connected through active
    relationships, so we can tell whether adding one more active
    relationship would close a cycle (path-compressed union-find)."""

    def __init__(self):
        self._parent: dict[str, str] = {}

    def find(self, x: str) -> str:
        self._parent.setdefault(x, x)
        root = x
        while self._parent[root] != root:
            root = self._parent[root]
        while self._parent[x] != root:
            self._parent[x], x = root, self._parent[x]
        return root

    def union(self, a: str, b: str) -> bool:
        """Returns True if a and b were in different components (and are
        now merged) — False if they were already connected, i.e. this edge
        would close a cycle."""
        ra, rb = self.find(a), self.find(b)
        if ra == rb:
            return False
        self._parent[ra] = rb
        return True


def _resolve_active_relationships(relationships: list[dict]) -> tuple[list[dict], list[str]]:
    """Guarantee the *active* relationship graph is a forest — i.e. there is
    at most one active filter path between any two tables — regardless of
    what the Excel's IsActive column says. This is what Fabric/Tabular
    actually requires; relying on a human to hand-mark every redundant leg
    of every cycle in the source Excel is exactly what left the earlier
    "ambiguous paths" bug reachable.

    Relationships explicitly marked inactive (IsActive=false in the Excel)
    stay inactive. Among the rest, a relationship is kept active only if it
    doesn't reconnect two tables that active relationships already connect
    (first relationship between any two components wins, in the order the
    Excel/auto-detection produced them); anything that would close a cycle
    is auto-deactivated — it's still created in the model, just inactive,
    so it remains usable in DAX via USERELATIONSHIP.
    """
    uf = _UnionFind()
    resolved: list[dict] = []
    auto_deactivated: list[str] = []
    for r in relationships:
        r = dict(r)
        if not r.get("is_active", True):
            resolved.append(r)
            continue
        if uf.union(r["from_table"], r["to_table"]):
            resolved.append(r)
        else:
            r["is_active"] = False
            auto_deactivated.append(
                f'{r["from_table"]}.{r["from_column"]} \u2192 {r["to_table"]}.{r["to_column"]}'
            )
            resolved.append(r)
    return resolved, auto_deactivated


def build_model_bim(
    server: str, database: str,
    tables_columns: dict[str, list[dict]],
    tables: list[dict],
    relationships: list[dict],
    measures: list[dict],
) -> tuple[dict, list[str]]:
    """Assemble a TMSL definition (model.bim) — DirectQuery against WH_Gold,
    one table per selected Excel row (using its live WH_Gold columns), plus
    the relationships and measures from the Excel.

    Returns (model_bim, auto_deactivated) — the second element lists any
    relationships that were forced inactive to keep the active-relationship
    graph a forest (see _resolve_active_relationships), so the caller can
    surface that in the job status/logs instead of it being silent."""

    relationships, auto_deactivated = _resolve_active_relationships(relationships)
    if auto_deactivated:
        logger.warning(
            "Auto-deactivated %d relationship(s) to avoid Fabric's 'ambiguous paths' "
            "error (multiple active filter paths between the same two tables): %s",
            len(auto_deactivated), "; ".join(auto_deactivated),
        )

    expr_lines = [
        "let",
        f'    Source = Sql.Database("{server}", "{database}")',
        "in",
        "    Source",
    ]

    model_tables = []
    for t in tables:
        key = f"{t['schema_name']}.{t['table_name']}"
        cols = tables_columns.get(key, [])
        table_measures = [
            {
                "name": m["measure_name"],
                "expression": m["expression"],
                **({"formatString": m["format_string"]} if m.get("format_string") else {}),
                "lineageTag": str(uuid.uuid4()),
            }
            for m in measures
            if m["schema_name"] == t["schema_name"] and m["table_name"] == t["table_name"]
        ]
        model_tables.append({
            "name": t["table_name"],
            "lineageTag": str(uuid.uuid4()),
            "columns": [
                {
                    "name": c["name"],
                    "dataType": c["tmsl_type"],
                    "sourceColumn": c["name"],
                    "lineageTag": str(uuid.uuid4()),
                }
                for c in cols
            ],
            **({"measures": table_measures} if table_measures else {}),
            "partitions": [
                {
                    "name": f"{t['table_name']}-partition",
                    "mode": "directQuery",
                    "source": {
                        "type": "m",
                        "expression": [
                            "let",
                            "    Source = DatabaseQuery,",
                            f'    Nav = Source{{[Schema="{t["schema_name"]}",Item="{t["table_name"]}"]}}[Data]',
                            "in",
                            "    Nav",
                        ],
                    },
                }
            ],
        })

    model_relationships = []
    for r in relationships:
        from_card, to_card = _CARDINALITY_MAP.get(r["cardinality"], ("many", "one"))
        model_relationships.append({
            "name": str(uuid.uuid4()),
            "fromTable": r["from_table"],
            "fromColumn": r["from_column"],
            "toTable": r["to_table"],
            "toColumn": r["to_column"],
            "fromCardinality": from_card,
            "toCardinality": to_card,
            "crossFilteringBehavior": (
                "bothDirections" if r["cross_filter_direction"] in ("both", "bidirectional") else "oneDirection"
            ),
            # TMSL relationships default to active if the key is omitted —
            # explicit here because leaving every relationship active is
            # exactly what produces Fabric's "ambiguous paths" provisioning
            # error whenever two tables are connected by more than one
            # active filter path (a schema with any relationship cycle
            # needs at least one side of the cycle marked inactive).
            "isActive": bool(r.get("is_active", True)),
        })

    return {
        "name": "Model",
        "compatibilityLevel": 1567,
        "model": {
            "culture": "en-US",
            "dataAccessOptions": {"legacyRedirects": True, "returnErrorValuesAsNull": True},
            "defaultPowerBIDataSourceVersion": "powerBI_V3",
            "sourceQueryCulture": "en-US",
            "expressions": [
                {
                    "name": "DatabaseQuery",
                    "kind": "m",
                    "expression": expr_lines,
                    "lineageTag": str(uuid.uuid4()),
                    "annotations": [{"name": "PBI_ResultType", "value": "Table"}],
                }
            ],
            "tables": model_tables,
            "relationships": model_relationships,
            "annotations": [{"name": "PBI_QueryOrder", "value": json.dumps([t["name"] for t in model_tables])}],
        },
    }, auto_deactivated


_DEFINITION_PBISM = {
    "version": "4.2",
    "settings": {},
}
_DOT_PLATFORM = {
    "$schema": "https://developer.microsoft.com/json-schemas/fabric/gitIntegration/platformProperties/2.0.0/schema.json",
    "metadata": {"type": "SemanticModel", "displayName": "SemanticModel"},
    "config": {"version": "2.0", "logicalId": str(uuid.uuid4())},
}


def _b64(obj) -> str:
    text = obj if isinstance(obj, str) else json.dumps(obj)
    return base64.b64encode(text.encode("utf-8")).decode("utf-8")


def build_definition_parts(model_bim: dict, display_name: str) -> list[dict]:
    platform = dict(_DOT_PLATFORM)
    platform["metadata"] = {"type": "SemanticModel", "displayName": display_name}
    return [
        {"path": "model.bim", "payload": _b64(model_bim), "payloadType": "InlineBase64"},
        {"path": "definition.pbism", "payload": _b64(_DEFINITION_PBISM), "payloadType": "InlineBase64"},
        {"path": ".platform", "payload": _b64(platform), "payloadType": "InlineBase64"},
    ]


# ── 4. Fabric REST API — create + poll ──────────────────────────────────


def _poll_lro(token: str, location_url: str) -> dict | None:
    headers = {"Authorization": f"Bearer {token}"}
    for _ in range(_LRO_MAX_POLLS):
        time.sleep(_LRO_POLL_INTERVAL)
        try:
            resp = httpx.get(location_url, headers=headers, timeout=_TIMEOUT)
            if resp.status_code != 200:
                continue
            data = resp.json()
            status = (data.get("status") or "").lower()
            if status in ("succeeded", "completed", "failed"):
                return data
        except Exception:
            continue
    return None


def find_semantic_model_by_name(token: str, workspace_id: str, display_name: str) -> dict | None:
    resp = httpx.get(
        f"{FABRIC_API_BASE}/workspaces/{workspace_id}/semanticModels",
        headers=_headers(token), timeout=_TIMEOUT,
    )
    if resp.status_code == 200:
        for item in resp.json().get("value", []):
            if item.get("displayName") == display_name:
                return item
    return None


def move_item_to_folder(token: str, workspace_id: str, item_id: str, folder_id: str) -> None:
    """Best-effort: move an existing item into the given folder. Used so
    that re-running a build also relocates a semantic model that was
    created before Gold-folder placement was wired in (or if it was ever
    dragged elsewhere), instead of silently leaving it wherever it already
    is. Never raises — folder placement is a nice-to-have, not something
    that should fail the whole build."""
    url = f"{FABRIC_API_BASE}/workspaces/{workspace_id}/items/{item_id}/move"
    try:
        resp = httpx.post(url, headers=_headers(token), json={"targetFolderId": folder_id}, timeout=_TIMEOUT)
        if not resp.is_success:
            logger.warning(f"Could not move semantic model to Gold folder ({resp.status_code}): {resp.text}")
    except Exception as e:
        logger.warning(f"Could not move semantic model to Gold folder: {e}")


def create_semantic_model(
    token: str, workspace_id: str, display_name: str, definition_parts: list[dict],
    folder_id: str | None = None,
) -> dict:
    """POST the definition to Fabric. Handles the 201 (immediate) and 202
    (long-running operation) cases, same pattern as notebook/pipeline
    upload. Returns the created (or existing, if it already exists) item.

    If `folder_id` is given (the Medallion Architecture / Gold folder),
    the item is created there directly; if it already exists somewhere
    else, it's moved into that folder as a best-effort follow-up."""
    url = f"{FABRIC_API_BASE}/workspaces/{workspace_id}/semanticModels"
    payload = {"displayName": display_name, "definition": {"parts": definition_parts}}
    if folder_id:
        payload["folderId"] = folder_id

    resp = httpx.post(url, headers=_headers(token), json=payload, timeout=_TIMEOUT)

    if resp.status_code == 201:
        return resp.json()

    if resp.status_code == 202:
        location = resp.headers.get("Location") or resp.headers.get("location")
        if location:
            lro_result = _poll_lro(token, location)
            if lro_result and (lro_result.get("status") or "").lower() == "failed":
                error_msg = (lro_result.get("error") or {}).get("message", "Semantic model provisioning failed")
                raise RuntimeError(f"Semantic model provisioning failed: {error_msg}")
        existing = find_semantic_model_by_name(token, workspace_id, display_name)
        if existing:
            return existing
        raise RuntimeError("Semantic model provisioning accepted but item not found afterwards")

    # 409 Conflict = an item with this name already exists. Re-running this
    # step should behave like the SQL deploys elsewhere in this module
    # (CREATE OR ALTER) — update the existing item's definition in place
    # rather than erroring, so "rebuild" after editing the Excel actually
    # picks up the new tables/relationships/measures.
    if resp.status_code == 409:
        existing = find_semantic_model_by_name(token, workspace_id, display_name)
        if existing and existing.get("id"):
            update_semantic_model_definition(token, workspace_id, existing["id"], definition_parts)
            if folder_id and existing.get("folderId") != folder_id:
                move_item_to_folder(token, workspace_id, existing["id"], folder_id)
            return existing
        if existing:
            return existing

    raise RuntimeError(f"Semantic model creation failed ({resp.status_code}): {resp.text}")


def update_semantic_model_definition(
    token: str, workspace_id: str, item_id: str, definition_parts: list[dict],
) -> None:
    """Overwrite an existing semantic model's definition (used when the
    display name already exists — rebuilding after an Excel re-upload)."""
    url = f"{FABRIC_API_BASE}/workspaces/{workspace_id}/semanticModels/{item_id}/updateDefinition"
    resp = httpx.post(
        url, headers=_headers(token), json={"definition": {"parts": definition_parts}}, timeout=_TIMEOUT
    )
    if resp.status_code == 202:
        location = resp.headers.get("Location") or resp.headers.get("location")
        if location:
            lro_result = _poll_lro(token, location)
            if lro_result and (lro_result.get("status") or "").lower() == "failed":
                error_msg = (lro_result.get("error") or {}).get("message", "Semantic model update failed")
                raise RuntimeError(f"Semantic model update failed: {error_msg}")
    elif not resp.is_success:
        raise RuntimeError(f"Semantic model update failed ({resp.status_code}): {resp.text}")