"""Business logic for the Fabric module.
 
Orchestrates repository calls and external Fabric API service calls.
"""
 
import asyncio
import logging
 
import httpx
from fastapi import HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
 
from app.modules.users.models.user import User
from app.modules.fabric import repository as repo
from app.modules.fabric.schema import (
    FabricCredentialCreate,
    LinkSourceConnectionRequest,
    MedallionConfigCreate,
    MetadataActionRequest,
    NotebookUploadRequest,
    PipelineRunRequest,
    PipelineUploadRequest,
    ProjectCreate,
    RunStatusUpdate,
    SourceConnectionCreate,
    WorkspaceProvisionRequest,
)
from app.modules.fabric.services.auth import get_fabric_token
from app.modules.fabric.services.notebook import list_workspace_notebooks
from app.modules.fabric.services.connection import (
    assign_role_to_connection,
    create_source_connection,
)
from app.modules.fabric.services.medallion import setup_medallion_architecture
from app.modules.fabric.services.metadata import setup_metadata_layer
from app.modules.fabric.services.pipeline import (
    build_replacements,
    get_pipeline_job_status,
    get_latest_item_job_status,
    list_local_pipelines,
    list_workspace_pipelines,
    run_fabric_pipeline,
    upload_pipelines,
    upload_itl_pipelines,
)
from app.modules.fabric.services.itl_config import read_otl_config, generate_itl_excel, parse_itl_excel, ensure_watermark_sp
from app.modules.fabric.services.notebook import list_local_notebooks, upload_notebooks, upload_itl_notebooks
from app.modules.fabric.services.workspace import provision_workspace
from app.modules.fabric.services.mapping_metadata import save_mapping_rows as _save_mapping_rows
from app.modules.fabric.services.mapping_metadata import read_latest_saved_mapping as _read_latest_saved_mapping
from app.core.config import settings
 
logger = logging.getLogger(__name__)

# asyncio.create_task() doesn't keep its task alive on its own — hold a
# strong reference here until each background deploy finishes.
_BACKGROUND_GOLD_SP_TASKS: set[asyncio.Task] = set()
_BACKGROUND_MASTER_SP_TASKS: set[asyncio.Task] = set()

# Serializes concurrent pipeline-deploy calls for the same project+connection.
# Without this, two overlapping deploy requests (e.g. the frontend's
# auto-deploy-after-notebook-create effect firing at nearly the same moment
# as a manual retry/click) can both pass the "does this pipeline already
# exist?" check before either has actually created it, and each independently
# create a duplicate DataPipeline item in Fabric.
_deploy_locks: dict[str, asyncio.Lock] = {}


def _get_deploy_lock(project_id: str, connection_name: str) -> asyncio.Lock:
    key = f"{project_id}:{connection_name}"
    lock = _deploy_locks.get(key)
    if lock is None:
        lock = asyncio.Lock()
        _deploy_locks[key] = lock
    return lock
 
 
# ── Helpers ──────────────────────────────────────────────────────────

def _ensure_job_registered(job_id: str, result: dict) -> None:
    """The 'View Mapping' resume flow hands the AI Mapping page a synthetic
    job_id (e.g. 'saved_<connection-id>') carrying a previously-saved
    result, so the summary can render without re-running the mapping. But
    every other AI Mapping action — Download CSV/Excel, Save to Metadata,
    manual-mapping overrides — is implemented as a lookup against the
    finin job_store by job_id, and that synthetic id was never written
    there, so those calls all 404'd with 'Job not ready'.

    Registering it here as a completed job (idempotent — a real, still-
    live job with this id is never overwritten) makes every job-based
    action work transparently for a resumed/saved mapping too.
    """
    from app.modules.finin.shared.job_store import get_job, create_job, update_job

    if get_job(job_id) is not None:
        return
    create_job(job_id)
    update_job(job_id, status="done", progress=100, total=100, message="Loaded saved mapping.", result=result)


async def _require_project(project_id: str, user: User, db: AsyncSession):
    project = await repo.get_project(db, project_id, str(user.id))
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project
 
 
async def _require_workspace(project_id: str, user: User, db: AsyncSession):
    project = await _require_project(project_id, user, db)
    if not project.workspace_id:
        raise HTTPException(
            status_code=400,
            detail="Provision a Fabric workspace for this project first",
        )
    return project
 
 
def _get_global_token():
    client_id = settings.FABRIC_CLIENT_ID
    client_secret = settings.FABRIC_CLIENT_SECRET
    tenant_id = settings.FABRIC_TENANT_ID
    if not all([client_id, client_secret, tenant_id]):
        raise HTTPException(
            status_code=500,
            detail="Fabric service principal not configured on server",
        )
    try:
        return get_fabric_token(client_id, client_secret, tenant_id)
    except RuntimeError as e:
        raise HTTPException(status_code=502, detail=str(e))
 
 
async def _get_project_token(project_id: str, db: AsyncSession):
    """Get Fabric token using stored project credentials; falls back to global env."""
    cred = await repo.get_fabric_credential(db, project_id)
    if cred:
        try:
            return get_fabric_token(cred.client_id, cred.client_secret, cred.tenant_id), cred
        except RuntimeError as e:
            raise HTTPException(status_code=502, detail=str(e))
    # Fallback to env-based global credentials
    return _get_global_token(), None
 
 
# ── Fabric Credentials ───────────────────────────────────────────────
 
 
async def save_fabric_credential_handler(
    project_id: str, payload: FabricCredentialCreate, user: User, db: AsyncSession
):
    """Save user-provided Fabric credentials for a project."""
    await _require_project(project_id, user, db)
 
    # Validate credentials by attempting to get a token
    try:
        get_fabric_token(payload.client_id, payload.client_secret, payload.tenant_id)
    except RuntimeError as e:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid credentials: {e}",
        )
 
    return await repo.save_fabric_credential(
        db,
        project_id=project_id,
        client_id=payload.client_id,
        client_secret=payload.client_secret,
        tenant_id=payload.tenant_id,
        capacity_id=payload.capacity_id,
        user_object_id=payload.user_object_id,
    )
 
 
async def get_fabric_credential_handler(
    project_id: str, user: User, db: AsyncSession
):
    """Get stored Fabric credentials for a project."""
    await _require_project(project_id, user, db)
    cred = await repo.get_fabric_credential(db, project_id)
    if not cred:
        raise HTTPException(status_code=404, detail="No Fabric credentials configured for this project")
    return cred
 
 
# ── Projects ─────────────────────────────────────────────────────────
 
 
async def list_projects(user: User, db: AsyncSession, app_type: str = "fabric"):
    return await repo.list_projects(db, str(user.id), app_type=app_type)


async def get_project_handler(project_id: str, user: User, db: AsyncSession):
    """Single-project lookup by id — lets the frontend check one project's
    workspace info without listing (and paying for) every project across
    both accelerators just to find the one it already has the id for.
    """
    project = await repo.get_project(db, project_id, str(user.id))
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project


async def create_project_handler(payload: ProjectCreate, user: User, db: AsyncSession):
    return await repo.create_project(
        db, name=payload.name, description=payload.description, user_id=str(user.id),
        app_type=payload.app_type,
    )
 
 
async def delete_project_handler(project_id: str, user: User, db: AsyncSession):
    project = await _require_project(project_id, user, db)
    await repo.soft_delete_project(db, project)
    return {"status": "deleted"}
 
 
# ── Source Connections (global per user) ─────────────────────────────
 
 
async def create_source_connection_handler(
    payload: SourceConnectionCreate, user: User, db: AsyncSession
):
    project = await repo.get_any_user_project_with_workspace(db, str(user.id))
    if not project:
        raise HTTPException(
            status_code=400,
            detail="Create a project with a provisioned workspace first",
        )
 
    # Infer db_type if not provided
    if not payload.db_type:
        server_lower = (payload.server or "").lower()
        if ".database.windows.net" in server_lower:
            payload.db_type = "Azure SQL"
        elif "oracle" in server_lower or ":1521" in server_lower:
            payload.db_type = "Oracle"
        elif "postgres" in server_lower or ":5432" in server_lower:
            payload.db_type = "PostgreSQL"
        elif "mysql" in server_lower or ":3306" in server_lower:
            payload.db_type = "MySQL"
        else:
            payload.db_type = "SQL Server"
 
    token, _ = await _get_project_token(project.id, db)

    db_username = payload.username
    db_password = payload.password
    if payload.auth_type == "ServicePrincipal":
        db_username = f"{payload.tenant_id}::{payload.client_id}"
        db_password = payload.client_secret or ""

    # Insert (and link) the row FIRST, before the slow Fabric API call below —
    # so the connection shows up as "creating" immediately, including across
    # a page reload that happens mid-request, instead of only appearing (or
    # silently vanishing) once this whole request finishes.
    record = await repo.create_source_connection_record(
        db,
        conn_name=payload.conn_name,
        db_type=payload.db_type,
        server=payload.server,
        database=payload.database,
        username=db_username,
        password=db_password,
        is_on_prem=payload.is_on_prem,
        gateway_name=payload.gateway_name,
        fabric_connection_id=None,
        user_id=str(user.id),
        status="creating",
    )
    if payload.project_id:
        try:
            await repo.create_project_link(
                db,
                project_id=payload.project_id,
                source_connection_id=record.id,
                connection_index=payload.connection_index,
            )
        except Exception:
            pass

    try:
        conn_data = create_source_connection(
            token=token,
            conn_name=payload.conn_name,
            db_type=payload.db_type,
            server=payload.server,
            database=payload.database,
            username=payload.username,
            password=payload.password,
            is_on_prem=payload.is_on_prem,
            gateway_name=payload.gateway_name,
            auth_type=payload.auth_type,
            tenant_id=payload.tenant_id,
            client_id=payload.client_id,
            client_secret=payload.client_secret,
        )
    except (RuntimeError, ValueError) as e:
        await repo.finalize_source_connection_status(db, record.id, status="failed", status_error=str(e))
        raise HTTPException(status_code=502, detail=str(e))
 
    fabric_conn_id = conn_data.get("id")
 
    if getattr(user, 'azure_oid', None):
        try:
            assign_role_to_connection(
                token=token,
                connection_id=fabric_conn_id,
                user_object_id=user.azure_oid,
                role="Owner",
            )
        except Exception:
            pass
 
    await repo.finalize_source_connection_status(db, record.id, status="active", fabric_connection_id=fabric_conn_id)
    await db.refresh(record)

    return record
 
 
async def list_source_connections(user: User, db: AsyncSession):
    return await repo.list_source_connections(db, str(user.id))
 
 
# ── Project ↔ Source Connection links ────────────────────────────────
 
 
async def link_source_connection(
    project_id: str,
    payload: LinkSourceConnectionRequest,
    user: User,
    db: AsyncSession,
):
    await _require_project(project_id, user, db)
    sc = await repo.get_source_connection(db, payload.source_connection_id, str(user.id))
    if not sc:
        raise HTTPException(status_code=404, detail="Source connection not found")
    return await repo.create_project_link(
        db,
        project_id=project_id,
        source_connection_id=payload.source_connection_id,
        connection_index=payload.connection_index,
    )
 
 
async def list_project_connections(project_id: str, user: User, db: AsyncSession):
    await _require_project(project_id, user, db)
    return await repo.list_project_links(db, project_id)
 
 
async def unlink_source_connection(
    project_id: str, link_id: str, user: User, db: AsyncSession
):
    await _require_project(project_id, user, db)
    link = await repo.get_project_link(db, link_id, project_id)
    if not link:
        raise HTTPException(status_code=404, detail="Link not found")
    await repo.delete_project_link(db, link)
    return {"status": "unlinked"}
 
 
# ── Medallion Architecture ──────────────────────────────────────────
 
 
async def create_medallion(
    project_id: str, payload: MedallionConfigCreate, user: User, db: AsyncSession
):
    project = await _require_workspace(project_id, user, db)
    token, _ = await _get_project_token(project_id, db)
 
    try:
        result = setup_medallion_architecture(
            token=token,
            workspace_id=project.workspace_id,
            bronze_is_lakehouse=payload.bronze_is_lakehouse,
            silver_is_lakehouse=payload.silver_is_lakehouse,
            gold_is_lakehouse=payload.gold_is_lakehouse,
            schema_enabled=payload.schema_enabled,
            bronze_name=payload.bronze_name,
            silver_name=payload.silver_name,
            gold_name=payload.gold_name,
        )
    except RuntimeError as e:
        raise HTTPException(status_code=502, detail=str(e))
    except httpx.HTTPStatusError as e:
        raise HTTPException(
            status_code=e.response.status_code,
            detail=f"Fabric API error: {e.response.status_code} – {e.response.text}",
        )
 
    return await repo.save_medallion_config(
        db,
        project_id,
        config_data=payload.model_dump(),
        result_data=result,
    )
 
 
async def list_medallion_configs(project_id: str, user: User, db: AsyncSession):
    await _require_project(project_id, user, db)
    return await repo.list_medallion_configs(db, project_id)
 
 
# ── Metadata / Log ──────────────────────────────────────────────────
 
 
async def create_metadata(
    project_id: str, payload: MetadataActionRequest, user: User, db: AsyncSession
):
    project = await _require_workspace(project_id, user, db)
    token, cred = await _get_project_token(project_id, db)
    _client_id = cred.client_id if cred else settings.FABRIC_CLIENT_ID
    _client_secret = cred.client_secret if cred else settings.FABRIC_CLIENT_SECRET
 
    try:
        result = setup_metadata_layer(
            token=token,
            workspace_id=project.workspace_id,
            client_id=_client_id,
            client_secret=_client_secret,
            action=payload.action,
        )
    except (RuntimeError, ValueError) as e:
        raise HTTPException(status_code=502, detail=str(e))
 
    # Persist metadata config
    if payload.action == "create_metadata":
        await repo.save_metadata_config(
            db,
            project_id,
            warehouse_id=result.get("warehouse_id"),
            metadata_created=True,
        )
    elif payload.action == "create_log":
        await repo.save_metadata_config(
            db, project_id, log_created=True
        )
 
    return result
 
 
# ── Notebooks ────────────────────────────────────────────────────────
 
 
async def upload_notebooks_handler(
    project_id: str, payload: NotebookUploadRequest, user: User, db: AsyncSession
):
    project = await _require_workspace(project_id, user, db)
    token, _ = await _get_project_token(project_id, db)
 
    # Resolve source connection id for status tracking
    sc = await repo.find_project_connection_by_name(db, project_id, payload.connection_name)
    source_connection_id = sc.id if sc else None
 
    # Auto-resolve db_type from source connection if not provided
    db_type = payload.db_type
    if not db_type and sc:
        db_type = sc.db_type or ""
 
    import asyncio
    try:
        results = await asyncio.to_thread(
            upload_notebooks,
            token=token,
            workspace_id=project.workspace_id,
            connection_name=payload.connection_name,
            connection_index=payload.connection_index,
            db_type=db_type,
            filenames=payload.filenames,
            app_mode=payload.app_mode,
        )
    except (RuntimeError, ValueError) as e:
        raise HTTPException(status_code=502, detail=str(e))
 
    # Persist upload status
    for r in results:
        await repo.save_config_upload(
            db,
            project_id=project_id,
            source_connection_id=source_connection_id,
            item_type="notebook",
            item_name=r["name"],
            status=r.get("status", "failed"),
            fabric_item_id=r.get("id"),
        )
 
    return {"status": "success", "results": results}
 
 
# ── Pipelines ────────────────────────────────────────────────────────
 
 
async def upload_pipelines_handler(
    project_id: str, payload: PipelineUploadRequest, user: User, db: AsyncSession
):
    project = await _require_workspace(project_id, user, db)
    token, _ = await _get_project_token(project_id, db)
 
    sc = await repo.find_project_connection_by_name(db, project_id, payload.connection_name)
    if not sc:
        raise HTTPException(
            status_code=404,
            detail=f"Connection '{payload.connection_name}' not linked to this project",
        )
 
    mc = await repo.get_medallion_config(db, project_id)
    if not mc:
        raise HTTPException(
            status_code=404, detail="Medallion config not found for this project"
        )
 
    # Auto-resolve db_type from source connection if not provided
    db_type = payload.db_type
    if not db_type:
        db_type = sc.db_type or ""
 
    conn_prefix = payload.connection_name + "_"
    _CONFIG_CREATION_SUFFIXES = (
        "01_NB_SQL_ConfigCreation", "01_NB_Oracle_ConfigCreation",
        "01_NB_Postgres_ConfigCreation", "01_NB_MySQL_ConfigCreation",
        "DBSource_ConfigCreation", "Oracle_ConfigCreation",
        "Config_Creation", "Config_Creation_modified", "SQLServer_ConfigCreation",
    )
 
    def _fetch_notebook_ids() -> dict[str, str]:
        """Look up the Fabric item IDs for this connection's ConfigCreation
        and BronzeToSilver notebooks, so pipeline JSON placeholders that
        reference them can be filled in.

        Retries for a bit if nothing turns up: a notebook that was *just*
        created (e.g. deploying its paired pipeline immediately afterwards,
        rather than as part of a separate later batch) isn't always visible
        yet via this list endpoint — Fabric's item listing lags item
        creation slightly. Without this retry, that race made pipeline
        deploys fail intermittently with an unresolved
        Replace_OTLMetaData_Creation_Id placeholder, especially right after
        a single notebook create+deploy (rather than a bulk batch, which
        had incidental delay built in).
        """
        import time as _time

        def _scan() -> dict[str, str]:
            ids: dict[str, str] = {}
            try:
                nb_url = f"https://api.fabric.microsoft.com/v1/workspaces/{project.workspace_id}/notebooks"
                nb_resp = httpx.get(nb_url, headers={"Authorization": f"Bearer {token}"}, timeout=httpx.Timeout(60.0, connect=10.0))
                if nb_resp.status_code == 200:
                    for item in nb_resp.json().get("value", []):
                        dn = item.get("displayName", "")
                        bare = dn[len(conn_prefix):] if dn.startswith(conn_prefix) else dn
                        if bare in _CONFIG_CREATION_SUFFIXES:
                            ids["Replace_OTLMetaData_Creation_Id"] = item["id"]
                        if bare in ("01_NB_BronzeToSilver", "NB_New_BronzeToSilver_DirectTable"):
                            ids["Replace_NB_New_BronzeToSilver_DirectTable_Id"] = item["id"]
            except Exception as e:
                logger.warning("Notebook lookup failed: %s", e)
            return ids

        ids = _scan()
        max_wait_seconds = 45
        interval_seconds = 3
        attempts = max_wait_seconds // interval_seconds
        for _ in range(attempts):
            if ids:
                break
            _time.sleep(interval_seconds)
            ids = _scan()
        return ids
 
    medallion_dict = {
        "bronze_item_id": mc.bronze_item_id,
        "silver_item_id": mc.silver_item_id,
        "gold_item_id": mc.gold_item_id,
    }
 
    import asyncio
    # Run notebook fetch and warehouse/endpoint fetch in parallel
    notebook_ids = await asyncio.to_thread(_fetch_notebook_ids)
 
    source_dict = {
        "db_type": sc.db_type,
        "database": sc.database,
        "fabric_connection_id": sc.fabric_connection_id,
        "conn_name": sc.conn_name,
        "notebook_ids": notebook_ids,
    }
 
    try:
        replacements = await asyncio.to_thread(
            build_replacements,
            token=token,
            workspace_id=project.workspace_id,
            source_connection=source_dict,
            medallion_config=medallion_dict,
        )
 
        # Fail early if critical Fabric references could not be resolved
        missing = []
        if not replacements.get("Replace_Bronze_Lakehouse_Id"):
            missing.append("Bronze Lakehouse ID (check Medallion config)")
        if not replacements.get("Replace_MetaData_Warehouse_Id"):
            missing.append("Metadata Warehouse ID (WH_MetaData not found)")
        if not replacements.get("Replace_MetaData_SQL_Endpoint"):
            missing.append("Metadata SQL Endpoint")
        if not replacements.get("Replace_Source_Connection_Id"):
            missing.append("Source Connection ID in Fabric")
        if missing:
            raise ValueError(
                f"Cannot upload pipelines – missing Fabric references: {'; '.join(missing)}. "
                "Ensure previous setup steps completed successfully and resources are fully provisioned."
            )
            


        async with _get_deploy_lock(project_id, payload.connection_name):
            results = await asyncio.to_thread(
                upload_pipelines,
                token=token,
                workspace_id=project.workspace_id,
                connection_name=payload.connection_name,
                connection_index=payload.connection_index,
                replacements=replacements,
                db_type=db_type,
                filenames=payload.filenames,
                app_mode=payload.app_mode,
            )
    except (RuntimeError, ValueError) as e:
        logger.error("Pipeline upload error: %s", e)
        raise HTTPException(status_code=502, detail=str(e))
 
    # Persist upload status
    for r in results:
        await repo.save_config_upload(
            db,
            project_id=project_id,
            source_connection_id=sc.id,
            item_type="pipeline",
            item_name=r["name"],
            status=r.get("status", "failed"),
            fabric_item_id=r.get("id"),
        )
 
    return {"status": "success", "results": results}
 
 
# ── Upload Status ────────────────────────────────────────────────────
 
 
async def get_upload_status_handler(
    project_id: str, user: User, db: AsyncSession, source_connection_id: str | None = None
):
    project = await repo.get_project(db, project_id, str(user.id))
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return await repo.get_config_uploads(db, project_id, source_connection_id)
 
 
# ── Pipeline Run / Status ────────────────────────────────────────────
 
 
async def list_workspace_pipelines_handler(
    project_id: str, user: User, db: AsyncSession
):
    project = await _require_workspace(project_id, user, db)
    token, _ = await _get_project_token(project_id, db)
    return list_workspace_pipelines(token, project.workspace_id)
 
 
async def upload_blob_config_handler(
    project_id: str,
    user: User,
    db: AsyncSession,
):
    from commands.discover_blob_structure import discover_blob_structure
    from app.modules.fabric.services.auth import get_onelake_token
    import json
    import httpx as _httpx
    import anyio

    project = await _require_workspace(project_id, user, db)
    mc = await repo.get_medallion_config(db, project_id)
    if not mc or not mc.bronze_item_id:
        raise HTTPException(status_code=400, detail="Bronze lakehouse config missing")

    # Find Azure Blob connection
    blob_conn = None
    links = await repo.list_project_links(db, project_id)
    for link in links:
        if link.source_connection and link.source_connection.db_type.lower() == "azure blob":
            blob_conn = link.source_connection
            break

    if not blob_conn:
        raise HTTPException(status_code=400, detail="No Azure Blob connection found for this project")

    # parse username format "tenant_id::client_id"
    parts = blob_conn.username.split("::") if blob_conn.username else []
    if len(parts) != 2:
        raise HTTPException(status_code=400, detail="Azure Blob connection missing valid Service Principal credentials")
    
    tenant_id, client_id = parts
    client_secret = blob_conn.password
    account_url = blob_conn.server
    container_name = blob_conn.database

    # Generate config
    try:
        config_data = await anyio.to_thread.run_sync(
            discover_blob_structure,
            tenant_id,
            client_id,
            client_secret,
            account_url,
            container_name
        )
    except Exception as e:
        logger.error("Blob structure discovery failed: %s", e)
        raise HTTPException(status_code=502, detail=f"Failed to discover blob structure: {e}")

    config_bytes = json.dumps(config_data, ensure_ascii=False).encode("utf-8")
    
    # Upload to OneLake
    cred = await repo.get_fabric_credential(db, project_id)
    if not cred or not cred.client_id or not cred.client_secret or not cred.tenant_id:
        raise HTTPException(status_code=400, detail="Missing fabric credentials")

    onelake_token = get_onelake_token(cred.client_id, cred.client_secret, cred.tenant_id)
    if not onelake_token:
        raise HTTPException(status_code=502, detail="Could not acquire OneLake token")

    bronze_lakehouse_id = mc.bronze_item_id
    dfs_base = (
        f"https://onelake.dfs.fabric.microsoft.com/"
        f"{project.workspace_id}/{bronze_lakehouse_id}/Files/config"
    )
    dfs_headers = {"Authorization": f"Bearer {onelake_token}"}
    
    # Ensure directory exists
    _httpx.put(f"{dfs_base}?resource=directory", headers=dfs_headers, timeout=30)
    
    file_name = "blob_config.json"
    dfs_file_url = f"{dfs_base}/{file_name}"
    
    # Create file
    cr = _httpx.put(f"{dfs_file_url}?resource=file", headers=dfs_headers, timeout=30)
    if cr.status_code not in (200, 201):
        raise HTTPException(status_code=502, detail=f"Failed to create config in OneLake: {cr.status_code} {cr.text}")

    # Append data
    _httpx.patch(
        f"{dfs_file_url}?action=append&position=0",
        headers={**dfs_headers, "Content-Type": "application/json", "Content-Length": str(len(config_bytes))},
        content=config_bytes,
        timeout=60,
    )
    # Flush data
    _httpx.patch(f"{dfs_file_url}?action=flush&position={len(config_bytes)}", headers=dfs_headers, timeout=30)
    
    logger.info("Uploaded blob_config.json to Bronze Lakehouse")

    # Record upload status in the DB
    from app.modules.fabric.models.config_upload import ConfigUpload
    from sqlalchemy import select
    
    # Check if entry already exists
    existing_upload = (
        await db.execute(
            select(ConfigUpload).where(
                ConfigUpload.project_id == project_id,
                ConfigUpload.item_name == "blob_config.json",
                ConfigUpload.item_type == "blob_config"
            )
        )
    ).scalar_one_or_none()

    if existing_upload:
        existing_upload.status = "success"
    else:
        new_upload = ConfigUpload(
            project_id=project_id,
            item_type="blob_config",
            item_name="blob_config.json",
            status="success",
        )
        db.add(new_upload)
    
    await db.commit()

    return {"status": "success", "message": "Blob config generated and uploaded to Bronze Lakehouse"}


async def run_pipeline_handler(
    project_id: str,
    pipeline_item_id: str,
    payload: PipelineRunRequest,
    user: User,
    db: AsyncSession,
):
    import asyncio

    project = await _require_workspace(project_id, user, db)
    token, _ = await _get_project_token(project_id, db)
 
    try:
        job_id = await asyncio.to_thread(
            run_fabric_pipeline,
            token=token,
            workspace_id=project.workspace_id,
            pipeline_item_id=pipeline_item_id,
            parameters=payload.parameters,
            job_type=payload.job_type,
        )
    except RuntimeError as e:
        raise HTTPException(status_code=502, detail=str(e))
 
    if not job_id:
        raise HTTPException(status_code=502, detail="Failed to start pipeline run")
 
    # Persist run status in config_uploads
    uploads = await repo.get_config_uploads(db, project_id)
    for u in uploads:
        if u.item_type == "pipeline" and u.fabric_item_id == pipeline_item_id:
            await repo.update_config_upload_run_status(
                db,
                project_id=project_id,
                item_type="pipeline",
                item_name=u.item_name,
                run_status="running",
                job_id=job_id,
            )
            break
 
    return {"job_id": job_id, "pipeline_id": pipeline_item_id, "status": "accepted"}
 
 
async def get_pipeline_job_status_handler(
    project_id: str,
    pipeline_item_id: str,
    job_id: str,
    user: User,
    db: AsyncSession,
):
    import asyncio

    project = await _require_workspace(project_id, user, db)
    token, _ = await _get_project_token(project_id, db)
 
    return await asyncio.to_thread(
        get_pipeline_job_status,
        token=token,
        workspace_id=project.workspace_id,
        pipeline_item_id=pipeline_item_id,
        job_id=job_id,
    )
 
 
async def get_latest_item_job_status_handler(
    project_id: str,
    pipeline_item_id: str,
    user: User,
    db: AsyncSession,
):
    """Latest job instance status for an item, regardless of who started it.

    Used for pipelines that are only ever invoked internally by a parent
    pipeline (e.g. an "Invoke Pipeline" activity) — we never have our own
    job_id for those, so get_pipeline_job_status_handler can't be used.
    """
    import asyncio

    project = await _require_workspace(project_id, user, db)
    token, _ = await _get_project_token(project_id, db)

    return await asyncio.to_thread(
        get_latest_item_job_status,
        token=token,
        workspace_id=project.workspace_id,
        pipeline_item_id=pipeline_item_id,
    )


async def update_run_status_handler(
    project_id: str,
    payload: RunStatusUpdate,
    user: User,
    db: AsyncSession,
):
    project = await repo.get_project(db, project_id, str(user.id))
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
 
    updated = await repo.update_config_upload_run_status(
        db,
        project_id=project_id,
        item_type=payload.item_type,
        item_name=payload.item_name,
        run_status=payload.run_status,
        job_id=payload.job_id,
    )
    if not updated:
        raise HTTPException(status_code=404, detail="Upload record not found")
    return updated
 
 
async def sync_pipeline_status_handler(
    project_id: str, user: User, db: AsyncSession
):
    """Check Fabric API for actual status of any 'running' pipelines and update DB."""
    project = await _require_workspace(project_id, user, db)
    token, _ = await _get_project_token(project_id, db)
 
    import asyncio

    uploads = await repo.get_config_uploads(db, project_id)
    running = [u for u in uploads if u.item_type == "pipeline" and u.run_status == "running" and u.job_id and u.fabric_item_id]
 
    pipelines = []
    updated = 0
    for u in running:
        try:
            status_data = await asyncio.to_thread(
                get_pipeline_job_status,
                token=token,
                workspace_id=project.workspace_id,
                pipeline_item_id=u.fabric_item_id,
                job_id=u.job_id,
            )
            new_status = status_data.get("status", "running")
            # "unknown" means the job instance couldn't be confirmed (e.g. a
            # transient 404 from Fabric) — report it as still running so we
            # don't flip a possibly-successful pipeline to "failed" in the UI.
            reported_status = "running" if new_status == "unknown" else new_status
            if new_status not in ("in_progress", "unknown", "running"):
                await repo.update_config_upload_run_status(
                    db,
                    project_id=project_id,
                    item_type="pipeline",
                    item_name=u.item_name,
                    run_status=new_status,
                    job_id=u.job_id,
                )
                updated += 1
            pipelines.append({"item_name": u.item_name, "new_status": reported_status})
        except Exception:
            pipelines.append({"item_name": u.item_name, "new_status": "running"})
 
    return {"updated": updated, "pipelines": pipelines}
 
 
# ── Workspace Provisioning ───────────────────────────────────────────
 
 
async def provision_workspace_handler(
    project_id: str,
    payload: WorkspaceProvisionRequest,
    user: User,
    db: AsyncSession,
):
    """Create a Fabric workspace, add the SP as Admin, and save workspace info on the project."""
    project = await _require_project(project_id, user, db)
 
    # Use stored project credentials if available, else fall back to env
    token, cred = await _get_project_token(project_id, db)
    capacity_id = cred.capacity_id if cred else settings.FABRIC_CAPACITY_ID
    user_object_id = (cred.user_object_id if cred else None) or getattr(user, 'azure_oid', None)
 
    try:
        result = provision_workspace(
            token=token,
            display_name=payload.workspace_name,
            capacity_id=capacity_id,
            user_azure_oid=user_object_id,
            user_fabric_token=payload.user_fabric_token,
            user_email=getattr(user, 'email', None),
        )
    except RuntimeError as e:
        logger.error("Workspace provisioning failed: %s", e)
        raise HTTPException(status_code=502, detail=str(e))
 
    # Save workspace_id and workspace_name on the project
    await repo.update_project_workspace(
        db, project,
        workspace_id=result["workspace_id"],
        workspace_name=result.get("workspace_name", payload.workspace_name),
        capacity_assigned=result.get("capacity_assigned", False),
    )
 
    # Update workspace_id on stored credentials
    if cred:
        await repo.update_credential_workspace_id(db, project_id, result["workspace_id"])
 
    return result
 
   
 
# ── ITL Pipelines Upload ─────────────────────────────────────────────
 
 
async def upload_itl_pipelines_handler(
    project_id: str, payload: PipelineUploadRequest, user: User, db: AsyncSession
):
    project = await _require_workspace(project_id, user, db)
    token, _ = await _get_project_token(project_id, db)
 
    sc = await repo.find_project_connection_by_name(db, project_id, payload.connection_name)
    if not sc:
        raise HTTPException(status_code=404, detail=f"Connection '{payload.connection_name}' not linked to this project")
 
    mc = await repo.get_medallion_config(db, project_id)
    if not mc:
        raise HTTPException(status_code=404, detail="Medallion config not found for this project")
 
    db_type = payload.db_type or sc.db_type or ""
    source_dict = {
        "db_type": sc.db_type,
        "database": sc.database,
        "fabric_connection_id": sc.fabric_connection_id,
        "conn_name": sc.conn_name,
        "notebook_ids": {},
    }
    medallion_dict = {
        "bronze_item_id": mc.bronze_item_id,
        "silver_item_id": mc.silver_item_id,
        "gold_item_id": mc.gold_item_id,
    }
 
    try:
        replacements = build_replacements(
            token=token,
            workspace_id=project.workspace_id,
            source_connection=source_dict,
            medallion_config=medallion_dict,
        )
        # Always use sc.conn_name (the exact stored name) so pipeline display names
        # and referenceName replacements are consistent in casing with OTL pipelines.
        results = upload_itl_pipelines(
            token=token,
            workspace_id=project.workspace_id,
            connection_name=sc.conn_name,
            connection_index=payload.connection_index,
            replacements=replacements,
            db_type=db_type,
        )
    except (RuntimeError, ValueError) as e:
        raise HTTPException(status_code=502, detail=str(e))
 
    for r in results:
        await repo.save_config_upload(
            db,
            project_id=project_id,
            source_connection_id=sc.id,
            item_type="pipeline",
            item_name=r["name"],
            status=r.get("status", "failed"),
            fabric_item_id=r.get("id"),
        )
 
    return {"status": "success", "results": results}
 
 
# ── ITL Config Download / Upload ─────────────────────────────────────
 
 
async def download_itl_config_handler(
    project_id: str, connection_name: str, user: User, db: AsyncSession
) -> bytes:
    project = await _require_workspace(project_id, user, db)
 
    sc = await repo.find_project_connection_by_name(db, project_id, connection_name)
    if not sc:
        raise HTTPException(status_code=404, detail=f"Connection '{connection_name}' not found")
 
    mc = await repo.get_metadata_config(db, project_id)
    if not mc:
        raise HTTPException(status_code=404, detail="Metadata config not found — run metadata setup first")
 
    try:
        from app.core.config import settings as app_settings
        from app.modules.fabric.services.metadata import get_warehouse_connection_string
 
        cred = await repo.get_fabric_credential(db, project_id)
        _client_id = cred.client_id if cred else app_settings.FABRIC_CLIENT_ID
        _client_secret = cred.client_secret if cred else app_settings.FABRIC_CLIENT_SECRET
 
        # Get token for Fabric API using project-specific credentials
        token, _ = await _get_project_token(project_id, db)
        if not mc.warehouse_id:
            raise HTTPException(status_code=400, detail="Warehouse ID not found in metadata config")
 
        # Get connection string from Fabric API
        warehouse_conn_str, _ = get_warehouse_connection_string(
            token, project.workspace_id, mc.warehouse_id
        )
 
        rows = read_otl_config(
            client_id=_client_id,
            client_secret=_client_secret,
            server=warehouse_conn_str,
            database=mc.warehouse_name or "WH_MetaData",
            config_schema_name=f"Config_{connection_name}",
            db_type=sc.db_type or "",
        )
        excel_bytes = generate_itl_excel(rows)
 
        # Mark as downloaded in DB (empty config_json = downloaded but not yet uploaded)
        import json as _json
        from app.modules.fabric.models.itl_watermark_config import ItlWatermarkConfig as _IWC
        existing = await repo.get_itl_watermark_config(db, project_id, connection_name)
        if not existing:
            db.add(_IWC(project_id=project_id, connection_name=connection_name, config_json="[]"))
            await db.commit()
 
        return excel_bytes
    except Exception as e:
        logger.error("ITL config download failed: %s", e)
        raise HTTPException(status_code=502, detail=f"Failed to read ITL config: {e}")
 
 
async def upload_itl_config_handler(
    project_id: str, connection_name: str, file_bytes: bytes, user: User, db: AsyncSession
):
    project = await _require_workspace(project_id, user, db)
 
    sc = await repo.find_project_connection_by_name(db, project_id, connection_name)
    if not sc:
        raise HTTPException(status_code=404, detail=f"Connection '{connection_name}' not found")
 
    try:
        rows = parse_itl_excel(file_bytes)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
 
    # ── Split Excel into two separate files and upload to OneLake DFS ────
    # File 1: TableSelection.xlsx  (sheet "Table_Selection")
    # File 2: {connection_name}.xlsx  (sheet "ITL_Config" / ConnectionName)
    token, _ = await _get_project_token(project_id, db)
    workspace_id = project.workspace_id
    medallion = await repo.get_medallion_config(db, project_id)
    onelake_path: str | None = None
 
    def _extract_sheet_bytes(source_bytes: bytes, sheet_name: str) -> bytes:
        """Extract a single sheet from an xlsx workbook and return it as a new xlsx file.
 
        For the ITL_Config sheet, columns D (CreatedWaterMarkField) and E (UpdatedWaterMarkField)
        are cross-sheet formulas referencing Table_Selection!E{n} / F{n}.
        openpyxl cannot evaluate these formulas, so we resolve them manually from the
        Table_Selection sheet before writing, ensuring values are stored as plain data.
        """
        import io as _io
        import openpyxl as _openpyxl
 
        src_wb = _openpyxl.load_workbook(_io.BytesIO(source_bytes), data_only=True)
        src_ws = src_wb[sheet_name]
 
        # Build a row->(created, updated) lookup from Table_Selection cols E & F
        # so we can resolve cross-sheet formula values for ITL_Config cols D & E.
        table_sel_values: dict = {}
        if sheet_name == "ITL_Config" and "Table_Selection" in src_wb.sheetnames:
            ts_ws = src_wb["Table_Selection"]
            for ts_row in ts_ws.iter_rows(min_row=2):
                row_num = ts_row[0].row
                created_val = ts_row[4].value if len(ts_row) > 4 else None
                updated_val = ts_row[5].value if len(ts_row) > 5 else None
                table_sel_values[row_num] = (created_val, updated_val)
 
        new_wb = _openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name
 
        # Copy column dimensions
        for col_letter, col_dim in src_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = col_dim.width
 
        # Copy row dimensions
        for row_idx, row_dim in src_ws.row_dimensions.items():
            new_ws.row_dimensions[row_idx].height = row_dim.height
 
        # Copy freeze panes
        if src_ws.freeze_panes:
            new_ws.freeze_panes = src_ws.freeze_panes
 
        # Copy cells with values and styles
        for row in src_ws.iter_rows():
            for cell in row:
                value = cell.value
 
                # Resolve cross-sheet formulas for ITL_Config cols D & E (columns 4 & 5)
                if sheet_name == "ITL_Config" and cell.row >= 2 and cell.column in (4, 5):
                    resolved = table_sel_values.get(cell.row)
                    if resolved is not None:
                        value = resolved[0] if cell.column == 4 else resolved[1]
                    elif isinstance(value, str) and value.startswith("="):
                        # Formula with no cached value — clear it
                        value = None
 
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=value)
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.alignment = cell.alignment.copy()
                    new_cell.number_format = cell.number_format
 
        buf = _io.BytesIO()
        new_wb.save(buf)
        return buf.getvalue()
 
    if medallion and medallion.bronze_item_id:
        bronze_lakehouse_id = medallion.bronze_item_id
 
        # OneLake DFS requires storage.azure.com audience — get a separate token
        from app.modules.fabric.services.auth import get_onelake_token as _get_onelake_token
        cred = await repo.get_fabric_credential(db, project_id)
        if cred:
            try:
                onelake_token = _get_onelake_token(cred.client_id, cred.client_secret, cred.tenant_id)
            except RuntimeError as e:
                logger.warning("Could not acquire OneLake token: %s", e)
                onelake_token = None
        else:
            onelake_token = None
 
        if onelake_token:
            import httpx as _httpx
            dfs_base = (
                f"https://onelake.dfs.fabric.microsoft.com/"
                f"{workspace_id}/{bronze_lakehouse_id}/Files/MetaData_ITL"
            )
            dfs_headers = {"Authorization": f"Bearer {onelake_token}"}
 
            # Ensure directory exists
            _httpx.put(f"{dfs_base}?resource=directory", headers=dfs_headers, timeout=30)
 
            def _dfs_upload(file_name: str, content: bytes) -> str:
                """Upload content to OneLake DFS and return the abfss path."""
                dfs_file_url = f"{dfs_base}/{file_name}"
                cr = _httpx.put(f"{dfs_file_url}?resource=file", headers=dfs_headers, timeout=30)
                if cr.status_code not in (200, 201):
                    raise RuntimeError(f"DFS create failed for {file_name}: {cr.status_code} {cr.text}")
                ar = _httpx.patch(
                    f"{dfs_file_url}?action=append&position=0",
                    headers={**dfs_headers, "Content-Type": "application/octet-stream", "Content-Length": str(len(content))},
                    content=content,
                    timeout=60,
                )
                if ar.status_code not in (200, 202):
                    raise RuntimeError(f"DFS append failed for {file_name}: {ar.status_code} {ar.text}")
                fr = _httpx.patch(f"{dfs_file_url}?action=flush&position={len(content)}", headers=dfs_headers, timeout=30)
                if fr.status_code not in (200, 202):
                    raise RuntimeError(f"DFS flush failed for {file_name}: {fr.status_code} {fr.text}")
                return (
                    f"abfss://{workspace_id}@onelake.dfs.fabric.microsoft.com/"
                    f"{bronze_lakehouse_id}/Files/MetaData_ITL/{file_name}"
                )
 
            try:
                # File 1: TableSelection.xlsx — contains the "Table_Selection" sheet
                sheet1_bytes = _extract_sheet_bytes(file_bytes, "Table_Selection")
                path1 = _dfs_upload("TableSelection.xlsx", sheet1_bytes)
                logger.info("Uploaded TableSelection.xlsx to OneLake: %s", path1)
 
                # File 2: {connection_name}.xlsx — contains the "ITL_Config" sheet
                sheet2_bytes = _extract_sheet_bytes(file_bytes, "ITL_Config")
                path2 = _dfs_upload(f"{connection_name}.xlsx", sheet2_bytes)
                logger.info("Uploaded %s.xlsx to OneLake: %s", connection_name, path2)
 
                # Store the connection-name file path as the primary reference
                onelake_path = path2
 
            except Exception as upload_err:
                logger.warning("OneLake DFS upload failed: %s", upload_err)
        else:
            logger.warning("Could not acquire OneLake token; skipping DFS upload")
 
    # Get warehouse connection and write to database
    # Save to local DB for reference (notebook handles writing to warehouse)
    import json
    from app.modules.fabric.models.itl_watermark_config import ItlWatermarkConfig
 
    existing = await repo.get_itl_watermark_config(db, project_id, connection_name)
    if existing:
        existing.config_json = json.dumps(rows)
        if onelake_path:
            existing.onelake_path = onelake_path
        db.add(existing)
    else:
        itl_config = ItlWatermarkConfig(
            project_id=project_id,
            connection_name=connection_name,
            config_json=json.dumps(rows),
        )
        if onelake_path:
            itl_config.onelake_path = onelake_path
        db.add(itl_config)
 
    await db.commit()
 
    # ── Ensure UpdateWaterMarkSP exists in this connection's Config_<name> schema ──
    try:
        from app.core.config import settings as app_settings
        from app.modules.fabric.services.metadata import get_warehouse_connection_string
 
        mc = await repo.get_metadata_config(db, project_id)
        cred = await repo.get_fabric_credential(db, project_id)
        _client_id = cred.client_id if cred else app_settings.FABRIC_CLIENT_ID
        _client_secret = cred.client_secret if cred else app_settings.FABRIC_CLIENT_SECRET
        if mc and mc.warehouse_id:
            warehouse_conn_str, _ = get_warehouse_connection_string(token, workspace_id, mc.warehouse_id)
            ensure_watermark_sp(
                client_id=_client_id,
                client_secret=_client_secret,
                server=warehouse_conn_str,
                database=mc.warehouse_name or "WH_MetaData",
                config_schema_name=f"Config_{connection_name}",
                app_mode=getattr(project, "app_type", "fabric") or "fabric",
            )
            logger.info("UpdateWaterMarkSP ensured in Config_%s", connection_name)
        else:
            logger.warning("Skipping UpdateWaterMarkSP creation: warehouse not found in metadata config")
    except Exception as sp_err:
        logger.warning("UpdateWaterMarkSP creation failed (non-fatal): %s", sp_err)
 
    # ── Upload ITL notebook to Fabric (so run endpoint can find it) ───
    try:
        db_type = sc.db_type if sc else "sql server"
        nb_results = upload_itl_notebooks(
            token=token,
            workspace_id=workspace_id,
            connection_name=connection_name,
            connection_index=1,
            db_type=db_type,
            filenames=["01_NB_IncrementalConfigCreation.ipynb"],
        )
        logger.info("ITL notebook upload results: %s", nb_results)
    except Exception as nb_err:
        logger.warning("ITL notebook upload failed (non-fatal): %s", nb_err)
 
    return {"status": "success", "rows_parsed": len(rows), "rows": rows, "onelake_path": onelake_path}
 
 
async def run_itl_notebook_handler(
    project_id: str,
    notebook_name: str,
    user: User,
    db: AsyncSession,
):
    """Run an ITL notebook in the Fabric workspace."""
    project = await _require_workspace(project_id, user, db)
    token, _ = await _get_project_token(project_id, db)
 
    from app.modules.fabric.services import notebook as nb_svc
 
    # Find the notebook item in workspace
    workspace_id = project.workspace_id
 
    # Notebooks are uploaded as "{connection_name}_01_NB_IncrementalConfigCreation".
    # Match by exact name, suffix, or substring of "IncrementalConfigCreation".
    items = list_workspace_notebooks(token, workspace_id)
    notebook_item = None
    for item in items:
        dn = item.get("displayName", "")
        if dn == notebook_name or dn.endswith(f"_{notebook_name}"):
            notebook_item = item
            break
    if not notebook_item:
        for item in items:
            if "IncrementalConfigCreation" in item.get("displayName", ""):
                notebook_item = item
                break
 
    if not notebook_item:
        available = [i.get("displayName") for i in items]
        raise HTTPException(status_code=404, detail=f"Notebook not found. Available: {available}")
 
    notebook_item_id = notebook_item.get("id")
    if not notebook_item_id:
        raise HTTPException(status_code=500, detail="Notebook ID not found")
 
    # Build parameters — pass onelake_path if available
    parameters: dict = {}
    # Fetch the most recently uploaded ITL config to get the onelake_path
    from app.modules.fabric import repository as _repo
    # Try to find any ITL watermark config for this project with an onelake_path
    from sqlalchemy import select as _select
    from app.modules.fabric.models.itl_watermark_config import ItlWatermarkConfig as _IWC
    result = await db.execute(
        _select(_IWC)
        .where(_IWC.project_id == project_id)
        .order_by(_IWC.created_at.desc())
    )
    latest_itl = result.scalars().first()
    if latest_itl and getattr(latest_itl, "onelake_path", None):
        parameters["metadata_path"] = {
            "value": latest_itl.onelake_path,
            "type": "string",
        }
        logger.info("Passing metadata_path to notebook: %s", latest_itl.onelake_path)
 
    # Run the notebook
    job_id = nb_svc.run_fabric_notebook(token, workspace_id, notebook_item_id, parameters or None)
    if not job_id:
        raise HTTPException(status_code=500, detail="Failed to start notebook job")
 
    # Persist run status so the UI can restore "running"/completed state after a refresh
    sc = await repo.find_project_connection_by_name(db, project_id, notebook_name.split("_IncrementalConfigCreation")[0].split("_01_NB")[0]) if "_" in notebook_name else None
    await repo.save_config_upload(
        db,
        project_id=project_id,
        source_connection_id=sc.id if sc else None,
        item_type="itl_notebook",
        item_name=notebook_item.get("displayName", notebook_name),
        status="success",
        fabric_item_id=notebook_item_id,
        run_status="running",
        job_id=job_id,
    )
 
    return {"job_id": job_id, "notebook_id": notebook_item_id, "status": "accepted"}
 
 
async def get_itl_notebook_status_handler(
    project_id: str,
    notebook_name: str,
    job_id: str,
    user: User,
    db: AsyncSession,
):
    """Get the status of an ITL notebook job."""
    project = await _require_workspace(project_id, user, db)
    token, _ = await _get_project_token(project_id, db)
 
    from app.modules.fabric.services import notebook as nb_svc
 
    workspace_id = project.workspace_id
 
    # Find the notebook item (same fuzzy match as run handler)
    items = list_workspace_notebooks(token, workspace_id)
    notebook_item = None
    for item in items:
        dn = item.get("displayName", "")
        if dn == notebook_name or dn.endswith(f"_{notebook_name}"):
            notebook_item = item
            break
    if not notebook_item:
        for item in items:
            if "IncrementalConfigCreation" in item.get("displayName", ""):
                notebook_item = item
                break
 
    if not notebook_item:
        raise HTTPException(status_code=404, detail=f"Notebook '{notebook_name}' not found")
 
    notebook_item_id = notebook_item.get("id")
 
    # Get job status
    status = nb_svc.get_notebook_job_status(token, workspace_id, notebook_item_id, job_id)
    raw_status = str(status.get("status", "")).lower()
    if raw_status in ("completed", "succeeded", "success"):
        await repo.update_config_upload_run_status(
            db, project_id=project_id, item_type="itl_notebook",
            item_name=notebook_item.get("displayName", notebook_name), run_status="success",
        )
    elif raw_status in ("failed", "cancelled", "canceled"):
        await repo.update_config_upload_run_status(
            db, project_id=project_id, item_type="itl_notebook",
            item_name=notebook_item.get("displayName", notebook_name), run_status="failed",
        )
    return status


async def resolve_project_fabric_connection(
    project_id: str, user: User, db: AsyncSession
) -> dict:
    """Resolve everything Finin needs to talk to this project's Fabric workspace
    directly over ODBC: the AAD service-principal creds and the workspace's SQL
    analytics endpoint (shared by every warehouse/lakehouse in that workspace --
    only the database name differs per item).
    """
    project = await _require_workspace(project_id, user, db)

    mc = await repo.get_metadata_config(db, project_id)
    if not mc or not mc.warehouse_id:
        raise HTTPException(status_code=404, detail="Metadata config not found - run metadata setup first")

    from app.core.config import settings as app_settings
    from app.modules.fabric.services.metadata import get_warehouse_connection_string

    cred = await repo.get_fabric_credential(db, project_id)
    client_id = cred.client_id if cred else app_settings.FABRIC_CLIENT_ID
    client_secret = cred.client_secret if cred else app_settings.FABRIC_CLIENT_SECRET
    tenant_id = cred.tenant_id if cred else app_settings.FABRIC_TENANT_ID

    token, _ = await _get_project_token(project_id, db)
    server, _ = get_warehouse_connection_string(token, project.workspace_id, mc.warehouse_id)

    # NOTE: Finin's Template schema is no longer read live from a Fabric
    # "Template_lakehouse" item — that item can live in any workspace and the
    # project's service principal frequently isn't granted access to it there
    # (18456 login failures). It's now cached locally in app.db and read via
    # mapping.repository.get_template_rows(); see mapping/router.py. No
    # cross-workspace Fabric lookup is needed here anymore.
    return {
        "server": server,
        "template_server": None,
        "client_id": client_id,
        "client_secret": client_secret,
        "tenant_id": tenant_id,
        "warehouse_name": mc.warehouse_name or "WH_MetaData",
    }


async def get_saved_finin_mapping_handler(
    project_id: str, connection_name: str, user: User, db: AsyncSession,
) -> dict | None:
    """Look up a previously-saved Finin mapping for this connection.

    Prefers the exact {stats, rows} JSON captured at save time
    (`ai_mapping_result_json`) — reconstructing it from
    [Config_<connection_name>].[SourceInformationSchemaMapped] instead is
    lossy, since that table only holds resolved template mappings plus
    "extension" (unmapped) source columns for the Bronze/Silver notebooks,
    not the original per-template-row match/unmatched status. Falls back to
    that warehouse reconstruction only for mappings saved before this column
    existed.
    """
    sc = await repo.find_project_connection_by_name(db, project_id, connection_name)
    if not sc or not sc.ai_mapping_saved:
        return None

    if sc.ai_mapping_result_json:
        import json

        try:
            result = json.loads(sc.ai_mapping_result_json)
            job_id = f"saved_{sc.id}"
            _ensure_job_registered(job_id, result)
            return {"job_id": job_id, "result": result}
        except (ValueError, TypeError):
            logger.warning("ai_mapping_result_json for connection %s was unreadable — "
                            "falling back to warehouse reconstruction.", connection_name)

    conn = await resolve_project_fabric_connection(project_id, user, db)

    import anyio

    reconstructed = await anyio.to_thread.run_sync(
        lambda: _read_latest_saved_mapping(
            client_id=conn["client_id"],
            client_secret=conn["client_secret"],
            server=conn["server"],
            database=conn["warehouse_name"],
            config_schema_name=f"Config_{connection_name}",
        )
    )
    if reconstructed:
        _ensure_job_registered(reconstructed["job_id"], reconstructed["result"])
    return reconstructed


async def save_finin_mapping_handler(
    project_id: str, connection_name: str, job_id: str, rows: list[dict], user: User, db: AsyncSession,
    progress_job_id: str | None = None, full_result: dict | None = None,
) -> int:
    """Persist Finin mapping results into [Config_<connection_name>].[SourceInformationSchemaMapped]."""
    sc = await repo.find_project_connection_by_name(db, project_id, connection_name)
    if not sc:
        raise HTTPException(status_code=404, detail=f"Connection '{connection_name}' not found")

    conn = await resolve_project_fabric_connection(project_id, user, db)

    import anyio

    on_progress = None
    if progress_job_id:
        from app.modules.finin.shared.job_store import update_job as _update_save_job

        def on_progress(done: int, total: int) -> None:  # noqa: E306
            pct = int((done / total) * 100) if total else 100
            _update_save_job(progress_job_id, progress=pct, message=f"Saved {done}/{total} rows")

    inserted = await anyio.to_thread.run_sync(
        lambda: _save_mapping_rows(
            client_id=conn["client_id"],
            client_secret=conn["client_secret"],
            server=conn["server"],
            database=conn["warehouse_name"],
            config_schema_name=f"Config_{connection_name}",
            job_id=job_id,
            rows=rows,
            on_progress=on_progress,
        )
    )

    # Persist the "already mapped" flag locally so the frontend can restore
    # this on page reload without re-querying the Fabric warehouse. Also
    # persist the exact result the live job computed (see
    # get_saved_finin_mapping_handler for why this is the source of truth
    # for the "View Mapping" resume view, rather than the warehouse table).
    sc.ai_mapping_saved = True
    if full_result is not None:
        import json

        sc.ai_mapping_result_json = json.dumps(full_result)
    db.add(sc)
    await db.commit()

    return inserted


async def start_deploy_gold_stored_procedures_handler(
    project_id: str, user: User, db: AsyncSession
) -> dict:
    """Kick off the WH_Gold stored-procedure deploy as a background job and
    return its job_id immediately, so the frontend can show live batch
    progress instead of blocking on one long request (89 procedures / ~90
    SQL batches was enough to feel like a hang with no feedback).
    """
    project = await _require_workspace(project_id, user, db)

    medallion = await repo.get_medallion_config(db, project_id)
    if not medallion or not medallion.gold_item_id:
        raise HTTPException(
            status_code=404,
            detail="Gold warehouse not found — complete the Medallion step first.",
        )

    cred = await repo.get_fabric_credential(db, project_id)
    if not cred:
        raise HTTPException(status_code=400, detail="Fabric credentials not configured for this project")

    mc = await repo.get_metadata_config(db, project_id)

    token, _ = await _get_project_token(project_id, db)

    from app.modules.fabric.services.metadata import get_warehouse_connection_string
    from app.modules.fabric.services import gold_stored_procedures as gsp
    from app.modules.finin.shared.job_store import create_job, update_job
    import anyio
    import uuid

    server, display_name = get_warehouse_connection_string(
        token, project.workspace_id, medallion.gold_item_id
    )
    database = display_name or medallion.gold_name or "WH_Gold"

    # Metadata (WH_MetaData) connection, resolved up front — best-effort:
    # if metadata config isn't set up, we still deploy the procedures, we
    # just skip recording them in Config_Gold.finin_gold_sp_details.
    meta_server = meta_database = None
    if mc and mc.warehouse_id:
        try:
            meta_server, meta_display_name = get_warehouse_connection_string(
                token, project.workspace_id, mc.warehouse_id
            )
            meta_database = meta_display_name or mc.warehouse_name or "WH_MetaData"
        except Exception as e:
            logger.warning(f"Could not resolve WH_MetaData connection for SP recording: {e}")

    job_id = f"gold_sp_{project_id}_{uuid.uuid4().hex[:8]}"
    total_batches = gsp.count_batches()
    create_job(job_id)
    update_job(job_id, status="running", total=total_batches, progress=0, message="Starting deployment…")

    async def _run():
        def on_progress(completed: int, total: int):
            update_job(job_id, progress=completed, total=total, message=f"Executing batch {completed}/{total}…")

        try:
            result = await anyio.to_thread.run_sync(
                lambda: gsp.deploy_stored_procedures(
                    client_id=cred.client_id,
                    client_secret=cred.client_secret,
                    server=server,
                    database=database,
                    on_progress=on_progress,
                )
            )
            result["database"] = database

            if meta_server and meta_database:
                try:
                    sp_names = gsp.extract_procedure_names()
                    recorded = await anyio.to_thread.run_sync(
                        lambda: gsp.record_sp_details(
                            client_id=cred.client_id,
                            client_secret=cred.client_secret,
                            server=meta_server,
                            database=meta_database,
                            sp_names=sp_names,
                        )
                    )
                    result["sp_details_recorded"] = recorded
                except Exception as e:
                    # Deployment itself succeeded — don't fail the whole job
                    # over the bookkeeping step; just surface it in the result.
                    logger.warning(f"Deployed SPs but failed to record Config_Gold.finin_gold_sp_details: {e}")
                    result["sp_details_recorded"] = 0
                    result["sp_details_error"] = str(e)

            update_job(job_id, status="done", progress=result["batches_executed"],
                       total=result["batches_executed"], message="Deployment complete", result=result)
        except Exception as e:
            update_job(job_id, status="failed", message=str(e))

    task = asyncio.create_task(_run())
    _BACKGROUND_GOLD_SP_TASKS.add(task)
    task.add_done_callback(_BACKGROUND_GOLD_SP_TASKS.discard)

    return {"status": "started", "job_id": job_id, "total": total_batches}

def get_deploy_gold_stored_procedures_status(job_id: str) -> dict:
    from app.modules.finin.shared.job_store import get_job
    job = get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return {
        "status": job.get("status", "unknown"),
        "progress": job.get("progress", 0),
        "total": job.get("total", 0),
        "message": job.get("message", ""),
        "result": job.get("result"),
    }


async def _resolve_gold_and_meta(project_id: str, user: User, db: AsyncSession):
    """Shared setup for both master-executer endpoints: workspace, Fabric
    credential, WH_Gold connection, and (best-effort) WH_MetaData connection."""
    project = await _require_workspace(project_id, user, db)

    medallion = await repo.get_medallion_config(db, project_id)
    if not medallion or not medallion.gold_item_id:
        raise HTTPException(
            status_code=404,
            detail="Gold warehouse not found — complete the Medallion step first.",
        )

    cred = await repo.get_fabric_credential(db, project_id)
    if not cred:
        raise HTTPException(status_code=400, detail="Fabric credentials not configured for this project")

    mc = await repo.get_metadata_config(db, project_id)
    token, _ = await _get_project_token(project_id, db)

    from app.modules.fabric.services.metadata import get_warehouse_connection_string

    gold_server, gold_display_name = get_warehouse_connection_string(
        token, project.workspace_id, medallion.gold_item_id
    )
    gold_database = gold_display_name or medallion.gold_name or "WH_Gold"

    meta_server = meta_database = None
    if mc and mc.warehouse_id:
        try:
            meta_server, meta_display_name = get_warehouse_connection_string(
                token, project.workspace_id, mc.warehouse_id
            )
            meta_database = meta_display_name or mc.warehouse_name or "WH_MetaData"
        except Exception as e:
            logger.warning(f"Could not resolve WH_MetaData connection: {e}")

    return cred, gold_server, gold_database, meta_server, meta_database


async def deploy_master_executer_handler(project_id: str, user: User, db: AsyncSession) -> dict:
    """Create [MasterExecuter].[sp_GoldExecute] (+ its schema/log table) in
    WH_Gold. Only 3 objects — fast enough to run synchronously, no job/
    progress bar needed for this part (only for actually *running* it)."""
    cred, gold_server, gold_database, _, _ = await _resolve_gold_and_meta(project_id, user, db)

    from app.modules.fabric.services import master_executer as mx
    import anyio

    result = await anyio.to_thread.run_sync(
        lambda: mx.deploy_master_executer(
            client_id=cred.client_id,
            client_secret=cred.client_secret,
            server=gold_server,
            database=gold_database,
        )
    )
    result["database"] = gold_database
    return result


async def start_execute_master_sp_handler(
    project_id: str, silver_lakehouse: str, user: User, db: AsyncSession
) -> dict:
    """Kick off EXEC [MasterExecuter].[sp_GoldExecute] as a background job.

    The procedure itself does all the work (loops over every active row in
    Config_Gold.finin_gold_sp_details and EXECs each [ims] procedure) — this
    just runs that single call on a background thread, while a second
    connection polls [MasterExecuter].[ExecutionLog] every couple seconds
    so the frontend gets a real "X of Y" progress bar instead of one long
    spinner for however many procedures are active.
    """
    cred, gold_server, gold_database, meta_server, meta_database = await _resolve_gold_and_meta(
        project_id, user, db
    )

    from app.modules.fabric.services import master_executer as mx
    from app.modules.finin.shared.job_store import create_job, update_job
    import anyio
    import uuid

    # Best-effort total for the progress bar — the run itself re-reads the
    # same table from inside SQL regardless of whether this succeeds.
    total = 0
    if meta_server and meta_database:
        total = await anyio.to_thread.run_sync(
            lambda: mx.get_active_sp_count(
                client_id=cred.client_id,
                client_secret=cred.client_secret,
                server=meta_server,
                database=meta_database,
            )
        )

    batch_id = mx.new_batch_id()
    job_id = f"master_sp_{project_id}_{uuid.uuid4().hex[:8]}"
    create_job(job_id)
    update_job(job_id, status="running", total=total, progress=0,
               message="Starting execution…", result={"batch_id": batch_id})

    async def _run():
        run_task = asyncio.create_task(
            anyio.to_thread.run_sync(
                lambda: mx.run_master_execute(
                    client_id=cred.client_id,
                    client_secret=cred.client_secret,
                    server=gold_server,
                    database=gold_database,
                    batch_id=batch_id,
                    silver_lakehouse=silver_lakehouse,
                )
            )
        )

        last_snapshot: dict = {"done": 0, "succeeded": 0, "failed": 0, "failed_names": []}
        while not run_task.done():
            await asyncio.sleep(2)
            try:
                last_snapshot = await anyio.to_thread.run_sync(
                    lambda: mx.poll_execution_log(
                        client_id=cred.client_id,
                        client_secret=cred.client_secret,
                        server=gold_server,
                        database=gold_database,
                        batch_id=batch_id,
                    )
                )
                update_job(
                    job_id,
                    progress=last_snapshot["done"],
                    total=max(total, last_snapshot["done"]),
                    message=f"Executed {last_snapshot['done']}/{max(total, last_snapshot['done']) or '?'} stored procedures…",
                )
            except Exception as e:
                # Transient polling error — the run itself keeps going
                # regardless; just skip this snapshot.
                logger.warning(f"Master SP progress poll failed: {e}")

        try:
            run_task.result()  # re-raise if run_master_execute failed
        except Exception as e:
            update_job(job_id, status="failed", message=str(e), result={"batch_id": batch_id, **last_snapshot})
            return

        # Final snapshot, now that the run has actually finished.
        try:
            last_snapshot = await anyio.to_thread.run_sync(
                lambda: mx.poll_execution_log(
                    client_id=cred.client_id,
                    client_secret=cred.client_secret,
                    server=gold_server,
                    database=gold_database,
                    batch_id=batch_id,
                )
            )
        except Exception as e:
            logger.warning(f"Master SP final progress poll failed: {e}")

        result = {"batch_id": batch_id, "database": gold_database, **last_snapshot}
        status = "failed" if last_snapshot["failed"] and last_snapshot["succeeded"] == 0 else "done"
        update_job(
            job_id, status=status,
            progress=last_snapshot["done"], total=max(total, last_snapshot["done"]),
            message="Execution complete", result=result,
        )

    task = asyncio.create_task(_run())
    _BACKGROUND_MASTER_SP_TASKS.add(task)
    task.add_done_callback(_BACKGROUND_MASTER_SP_TASKS.discard)

    return {"status": "started", "job_id": job_id, "total": total, "batch_id": batch_id}


def get_execute_master_sp_status(job_id: str) -> dict:
    from app.modules.finin.shared.job_store import get_job
    job = get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return {
        "status": job.get("status", "unknown"),
        "progress": job.get("progress", 0),
        "total": job.get("total", 0),
        "message": job.get("message", ""),
        "result": job.get("result"),
    }


# ── Semantic Model (Finin) ───────────────────────────────────────────

_BACKGROUND_SEMANTIC_MODEL_TASKS: set[asyncio.Task] = set()

# Item name used in Fabric and as the ConfigUpload item_name key — one
# semantic model per project for now (matches the "one WH_Gold, hard-coded"
# scope described in semantic_model.py).
_SEMANTIC_MODEL_ITEM_TYPE = "semantic_model"

# TESTING ONLY: the create-workspace -> load-data-to-WH_Gold flow hasn't
# been run for every project yet, but one workspace (sm.HARDCODED_*) already
# has real data loaded into its WH_Gold. Point the semantic model builder
# there directly for now instead of resolving each project's own (likely
# still-empty) Gold warehouse via _resolve_gold_and_meta().
# TODO: flip this back to False once each project loads its own WH_Gold —
# the dynamic per-project resolution path is already implemented below,
# just bypassed while this is True.
_USE_HARDCODED_TEST_GOLD = True


async def upload_semantic_model_excel_handler(
    project_id: str, file_bytes: bytes, filename: str, user: User, db: AsyncSession
) -> dict:
    """Parse the uploaded Tables/Relationships/Measures Excel and persist it
    against the project — does NOT touch Fabric or WH_Gold; that happens in
    start_build_semantic_model_handler."""
    await _require_project(project_id, user, db)

    from app.modules.fabric.services import semantic_model as sm

    try:
        parsed = sm.parse_semantic_excel(file_bytes)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    record = await repo.save_semantic_model_upload(
        db,
        project_id=project_id,
        filename=filename,
        tables=parsed["tables"],
        relationships=parsed["relationships"],
        measures=parsed["measures"],
    )
    return {
        "filename": record.filename,
        "tables_count": len(parsed["tables"]),
        "relationships_count": len(parsed["relationships"]),
        "measures_count": len(parsed["measures"]),
        "uploaded_at": record.updated_at,
    }


async def get_semantic_model_status_handler(project_id: str, user: User, db: AsyncSession) -> dict:
    """Restore Config-page state after a reload: what's been uploaded, and
    the last known build status/result."""
    await _require_project(project_id, user, db)
    import json

    upload = await repo.get_semantic_model_upload(db, project_id)
    excel = None
    if upload:
        excel = {
            "filename": upload.filename,
            "tables_count": len(json.loads(upload.tables_json)),
            "relationships_count": len(json.loads(upload.relationships_json)),
            "measures_count": len(json.loads(upload.measures_json)),
            "uploaded_at": upload.updated_at,
        }

    uploads = await repo.get_config_uploads(db, project_id)
    build_row = next((u for u in uploads if u.item_type == _SEMANTIC_MODEL_ITEM_TYPE), None)
    build = None
    if build_row:
        build = {
            "status": build_row.status,
            "fabric_item_id": build_row.fabric_item_id,
            "job_id": build_row.job_id,
            "display_name": build_row.item_name,
        }

    return {"excel": excel, "build": build}


async def start_build_semantic_model_handler(
    project_id: str, user: User, db: AsyncSession
) -> dict:
    """Kick off semantic-model creation as a background job: read the
    saved Excel, auto-detect relationships/measures from WH_Gold when the
    Excel didn't supply them, pull live column info for each selected
    table, build a TMSL definition, and create/update the Semantic Model
    item inside the project's own Gold folder (01_Medallion Architecture /
    Gold) in Fabric — polling the long-running operation under the hood.
    Returns immediately with a job_id to poll.
    """
    upload = await repo.get_semantic_model_upload(db, project_id)
    if not upload:
        raise HTTPException(
            status_code=400,
            detail="Upload the Tables/Relationships/Measures Excel first.",
        )

    project = await _require_workspace(project_id, user, db)

    from app.modules.fabric.services import semantic_model as sm
    from app.modules.fabric.services.medallion import get_folder_id
    from app.modules.finin.shared.job_store import create_job, update_job
    import anyio
    import json
    import uuid as _uuid

    if _USE_HARDCODED_TEST_GOLD:
        # TESTING: use this project's own Fabric credential (still needed
        # to authenticate), but point the connection + item placement at
        # the one workspace that already has data loaded, instead of this
        # project's own WH_Gold.
        cred = await repo.get_fabric_credential(db, project_id)
        if not cred:
            raise HTTPException(status_code=400, detail="Fabric credentials not configured for this project")
        gold_workspace_id = sm.HARDCODED_WORKSPACE_ID
        gold_server = sm.HARDCODED_GOLD_SERVER
        gold_database = sm.HARDCODED_GOLD_DATABASE
        token, _ = await _get_project_token(project_id, db)
    else:
        # Resolve the calling project's own workspace + WH_Gold connection
        # (same helper Master SP uses).
        cred, gold_server, gold_database, _, _ = await _resolve_gold_and_meta(project_id, user, db)
        gold_workspace_id = project.workspace_id
        token, _ = await _get_project_token(project_id, db)

    tables = json.loads(upload.tables_json)
    relationships = json.loads(upload.relationships_json)
    measures = json.loads(upload.measures_json)
    display_name = f"{project.name}_SemanticModel".replace(" ", "_")

    # Land the model in the same place the Gold warehouse itself lives:
    # 01_Medallion Architecture / Gold, instead of the workspace root.
    med_folder_id = await anyio.to_thread.run_sync(
        lambda: get_folder_id(token, gold_workspace_id, "01_Medallion Architecture")
    )
    gold_folder_id = None
    if med_folder_id:
        gold_folder_id = await anyio.to_thread.run_sync(
            lambda: get_folder_id(token, gold_workspace_id, "Gold", med_folder_id)
        )
    if not gold_folder_id:
        logger.warning(
            f"Could not find the Gold folder under 01_Medallion Architecture in workspace "
            f"{gold_workspace_id}; run the Medallion step there first. Semantic model will be "
            "created at the workspace root instead."
        )

    job_id = f"semantic_model_{project_id}_{_uuid.uuid4().hex[:8]}"
    create_job(job_id)
    update_job(job_id, status="running", total=len(tables), progress=0, message="Starting…")

    async def _run():
        # This runs after the request that created it has already returned
        # its response, so it can't reuse the request-scoped `db` session
        # (which FastAPI tears down once the request completes) — open a
        # fresh one here instead, same as any other detached background job
        # would need to.
        from app.db.session import async_session_maker

        try:
            update_job(job_id, message="Detecting relationships and measures from WH_Gold…", progress=0)
            try:
                resolved_tables, resolved_relationships, resolved_measures = await anyio.to_thread.run_sync(
                    lambda: sm.auto_detect_relationships_and_measures(
                        client_id=cred.client_id,
                        client_secret=cred.client_secret,
                        server=gold_server,
                        database=gold_database,
                        tables=tables,
                        existing_relationships=relationships,
                        existing_measures=measures,
                    )
                )
            except Exception as e:
                # Auto-detection is a best-effort enhancement, not a hard
                # requirement — if it blows up for any reason (a table
                # that INFORMATION_SCHEMA sees but SELECT can't reach,
                # missing permissions, etc.) fall back to exactly what the
                # Excel provided rather than failing the whole build.
                logger.warning(f"Relationship/measure auto-detection failed, falling back to Excel-only: {e}")
                resolved_tables, resolved_relationships, resolved_measures = tables, relationships, measures

            # DIAGNOSTIC: confirms whether this run is actually using the
            # freshly-uploaded relationships (with IsActive respected) or
            # stale data — check this log line if an "ambiguous paths"
            # error recurs after re-uploading the Excel.
            _inactive_ct = sum(1 for r in resolved_relationships if not r.get("is_active", True))
            logger.info(
                f"[semantic_model] {len(resolved_relationships)} relationship(s) resolved, "
                f"{_inactive_ct} marked inactive"
            )

            update_job(
                job_id,
                message=f"Reading {len(resolved_tables)} table schema(s) from WH_Gold…",
                progress=0,
            )
            tables_columns = await anyio.to_thread.run_sync(
                lambda: sm.fetch_table_columns(
                    client_id=cred.client_id,
                    client_secret=cred.client_secret,
                    server=gold_server,
                    database=gold_database,
                    tables=resolved_tables,
                )
            )
            update_job(job_id, message="Building semantic model definition…",
                       progress=len(resolved_tables) // 2)

            model_bim = sm.build_model_bim(
                server=gold_server,
                database=gold_database,
                tables_columns=tables_columns,
                tables=resolved_tables,
                relationships=resolved_relationships,
                measures=resolved_measures,
            )
            parts = sm.build_definition_parts(model_bim, display_name)

            update_job(job_id, message="Creating semantic model in Fabric…", progress=len(resolved_tables))
            item = await anyio.to_thread.run_sync(
                lambda: sm.create_semantic_model(
                    token=token,
                    workspace_id=gold_workspace_id,
                    display_name=display_name,
                    definition_parts=parts,
                    folder_id=gold_folder_id,
                )
            )

            async with async_session_maker() as bg_db:
                await repo.save_config_upload(
                    bg_db,
                    project_id=project_id,
                    source_connection_id=None,
                    item_type=_SEMANTIC_MODEL_ITEM_TYPE,
                    item_name=display_name,
                    status="success",
                    fabric_item_id=item.get("id"),
                    job_id=job_id,
                )

            result = {
                "display_name": display_name,
                "fabric_item_id": item.get("id"),
                "workspace_id": gold_workspace_id,
                "folder": "01_Medallion Architecture/Gold" if gold_folder_id else "workspace root",
                "tables": len(resolved_tables),
                "relationships": len(resolved_relationships),
                "measures": len(resolved_measures),
            }
            update_job(job_id, status="done", progress=len(resolved_tables), total=len(resolved_tables),
                       message="Semantic model created", result=result)
        except Exception as e:
            logger.exception("Semantic model build failed")
            try:
                async with async_session_maker() as bg_db:
                    await repo.save_config_upload(
                        bg_db,
                        project_id=project_id,
                        source_connection_id=None,
                        item_type=_SEMANTIC_MODEL_ITEM_TYPE,
                        item_name=display_name,
                        status="failed",
                        job_id=job_id,
                    )
            except Exception:
                logger.exception("Also failed to record semantic-model failure status")
            update_job(job_id, status="failed", message=str(e))

    task = asyncio.create_task(_run())
    _BACKGROUND_SEMANTIC_MODEL_TASKS.add(task)
    task.add_done_callback(_BACKGROUND_SEMANTIC_MODEL_TASKS.discard)

    return {"status": "started", "job_id": job_id, "total": len(tables)}


def get_build_semantic_model_status(job_id: str) -> dict:
    from app.modules.finin.shared.job_store import get_job
    job = get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return {
        "status": job.get("status", "unknown"),
        "progress": job.get("progress", 0),
        "total": job.get("total", 0),
        "message": job.get("message", ""),
        "result": job.get("result"),
    }


# ── Blob Structure Discovery ────────────────────────────────────────


async def discover_blob_structure_handler(
    prefix: str = "Data/",
    archive_root: str = "Archive/",
) -> dict:
    """
    Discover the folder structure in Azure Blob Storage under the given prefix
    and return the generated config dict.

    Uses BLOB_* environment variables for authentication.
    """
    import anyio

    # Import and run the blocking discovery logic in a thread
    from commands.discover_blob_structure import discover_blob_structure

    try:
        config = await anyio.to_thread.run_sync(
            lambda: discover_blob_structure(prefix=prefix, archive_root=archive_root)
        )
    except Exception as e:
        logger.error("Blob structure discovery failed: %s", e)
        raise HTTPException(
            status_code=502,
            detail=f"Failed to discover blob structure: {e}",
        )

    return config