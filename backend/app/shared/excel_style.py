"""Shared Excel (openpyxl) styling used across every generated .xlsx in the
app, so ITL config sheets, Finin mapping exports, etc. all look the same:
bold white-on-blue header row, alternating light-blue/white row banding, and
a soft-yellow highlight for "suggested"/derived columns.

Mirrors the styling originally written for backend/app/modules/fabric/services/
itl_config.py — kept here so it isn't duplicated per-module.
"""
from __future__ import annotations

CLR_HEADER_DEFAULT = "4472C4"   # blue
CLR_HEADER_ALT = "70AD47"       # green (for a second/related sheet)
CLR_ROW_ODD = "EBF3FB"          # light blue
CLR_ROW_EVEN = "FFFFFF"         # white
CLR_SUGGESTED = "FFF2CC"        # soft yellow — suggested/derived field
CLR_HEADER_FONT = "FFFFFF"
CLR_BORDER = "BDD7EE"


def _apply_border(cell):
    from openpyxl.styles import Border, Side
    thin = Side(style="thin", color=CLR_BORDER)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header_cell(cell, hex_color: str = CLR_HEADER_DEFAULT):
    from openpyxl.styles import PatternFill, Font, Alignment
    cell.fill = PatternFill("solid", fgColor=hex_color)
    cell.font = Font(bold=True, color=CLR_HEADER_FONT, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _apply_border(cell)


def style_row_cell(cell, row_idx: int, highlight: bool = False):
    from openpyxl.styles import PatternFill, Alignment
    color = CLR_SUGGESTED if highlight else (CLR_ROW_ODD if row_idx % 2 == 1 else CLR_ROW_EVEN)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    _apply_border(cell)


def style_worksheet(
    ws,
    num_data_rows: int,
    num_cols: int,
    header_color: str = CLR_HEADER_DEFAULT,
    highlight_col_indices: set[int] | None = None,
    col_widths: list[int] | None = None,
    header_row: int = 1,
) -> None:
    """Style an existing worksheet (e.g. one written by ``df.to_excel``) in
    place: bold colored header row, banded data rows, optional per-column
    highlight, freeze panes below the header, and auto column widths.

    `highlight_col_indices` is a set of 1-based column numbers to always
    render with the "suggested" yellow fill instead of the row's normal band
    color — use for derived/suggested columns like watermark fields.
    """
    ws.row_dimensions[header_row].height = 30
    highlight_col_indices = highlight_col_indices or set()

    for ci in range(1, num_cols + 1):
        cell = ws.cell(row=header_row, column=ci)
        style_header_cell(cell, header_color)
        if col_widths and ci <= len(col_widths):
            ws.column_dimensions[cell.column_letter].width = col_widths[ci - 1]
        else:
            ws.column_dimensions[cell.column_letter].width = 22

    for ri in range(1, num_data_rows + 1):
        excel_row = header_row + ri
        for ci in range(1, num_cols + 1):
            cell = ws.cell(row=excel_row, column=ci)
            style_row_cell(cell, ri, highlight=(ci in highlight_col_indices))

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate